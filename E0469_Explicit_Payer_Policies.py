#!/usr/bin/env python3
"""
Generate E0469 EXPLICIT Payer Coverage Analysis Spreadsheet
HCPCS E0469: Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device
Effective: October 1, 2024

THIS FILE ONLY CONTAINS PAYERS THAT EXPLICITLY MENTION E0469 IN THEIR POLICIES
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "E0469 Payer Coverage"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
covered_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
not_covered_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
investigational_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
case_by_case_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Payer Name",
    "Payer Type",
    "Coverage Status",
    "Prior Auth Required",
    "Investigational/Experimental",
    "Not Medically Necessary",
    "Policy/Effective Date",
    "Policy Number",
    "Notes/Details",
    "Source URL"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# ONLY PAYERS WITH EXPLICIT E0469 MENTION IN POLICY
payer_data = [
    # CMS/Medicare - E0469 added to fee schedule 10/1/2024
    {
        "name": "CMS DMEPOS Fee Schedule",
        "type": "Medicare/CMS",
        "coverage": "Case-by-Case (No LCD)",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "10/01/2024",
        "policy_num": "No LCD/NCD",
        "notes": "E0469 added to DMEPOS fee schedule effective 10/1/2024. Category codes 01, 60. No specific LCD - claims reviewed individually. Capped rental with RR modifier.",
        "source": "https://www.cms.gov/files/document/r12835cp.pdf"
    },
    # Noridian Medicare JA DME - E0469 billing guidance
    {
        "name": "Noridian Medicare (JA DME)",
        "type": "Medicare MAC",
        "coverage": "Case-by-Case (No LCD)",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "10/01/2024",
        "policy_num": "Article 6547796",
        "notes": "E0469 billing article published. No specific LCD policy exists. Claims reviewed individually. Must document medical necessity with valid order, proof of delivery, and medical records.",
        "source": "https://med.noridianmedicare.com/web/jadme/article-detail/-/view/6547796/lung-expansion-airway-clearance-continuous-high-frequency-oscillation-and-nebulization-device-hcpcs-code-e0469"
    },
    # CGS Medicare JB DME - E0469 billing guidance
    {
        "name": "CGS Medicare (JB DME)",
        "type": "Medicare MAC",
        "coverage": "Case-by-Case (No LCD)",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "01/07/2025",
        "policy_num": "COPE181758",
        "notes": "E0469 correct coding and billing article published. No LCD policy exists. Claims reviewed individually. Capped rental item billed with RR modifier plus KH/KI/KJ.",
        "source": "https://www.cgsmedicare.com/jb/pubs/news/2025/07/cope181758.html"
    },
    # Medica - E0469 EXPLICITLY NOT COVERED
    {
        "name": "Medica",
        "type": "Regional",
        "coverage": "NOT COVERED",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "Volara OLE Coverage Policy",
        "notes": "E0469 EXPLICITLY listed. Oscillating lung expansion systems (Volara, BiWaze Clear) considered INVESTIGATIVE and UNPROVEN - therefore NOT COVERED. Insufficient peer-reviewed medical literature.",
        "source": "https://partner.medica.com/-/media/documents/provider/coverage-policies/volara-oscillation-and-lung-expansion-cp.pdf"
    },
    # HealthPartners - E0469 INVESTIGATIONAL
    {
        "name": "HealthPartners",
        "type": "Regional",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "Airway Clearance Policy",
        "notes": "E0469 and A7021 EXPLICITLY listed as 'investigational/experimental for home use.' Prior auth not applicable because devices considered investigational. Provider/facility liable unless member signs waiver.",
        "source": "https://www.healthpartners.com/ucm/groups/public/@hp/@public/@cc/documents/documents/aentry_045636.pdf"
    },
    # Minnesota MHCP Medicaid - E0469 NOT COVERED
    {
        "name": "Minnesota MHCP (Medicaid)",
        "type": "Medicaid",
        "coverage": "NOT COVERED",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "DHS16_152808",
        "notes": "E0469 EXPLICITLY NOT COVERED for any indication. Lung expansion airway clearance devices not considered standard of care in community setting. Substantive research lacking.",
        "source": "https://www.dhs.state.mn.us/main/idcplg?IdcService=GET_DYNAMIC_CONVERSION&dDocName=DHS16_152808"
    },
    # BCBS Florida - E0469 INVESTIGATIONAL
    {
        "name": "Blue Cross Blue Shield Florida",
        "type": "BCBS",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "09/25/2025 (reviewed)",
        "policy_num": "09-E0000-28",
        "notes": "E0469 EXPLICITLY mentioned. Volara System OLE therapy device 'considered experimental or investigational.' Insufficient published clinical data for health outcomes conclusions.",
        "source": "https://mcgs.bcbsfl.com/MCG?mcgId=09-E0000-28&pv=false"
    },
    # Premera Blue Cross - E0469 INVESTIGATIONAL
    {
        "name": "Premera Blue Cross",
        "type": "BCBS",
        "coverage": "Investigational",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "10/01/2024 (codes added)",
        "policy_num": "1.01.539",
        "notes": "E0469 and A7021 EXPLICITLY added effective 10/1/2024. OLE devices (Volara, MetaNeb 4, BiWaze Clear) 'considered investigational' for treatment of respiratory conditions.",
        "source": "https://www.premera.com/medicalpolicies-individual/1.01.539.pdf"
    },
    # Kaiser WA Medicare - E0469 COVERED
    {
        "name": "Kaiser Permanente WA (Medicare)",
        "type": "Medicare Advantage",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "No",
        "not_med_necessary": "No",
        "date": "07/17/2025",
        "policy_num": "HFCWO Criteria",
        "notes": "E0469 EXPLICITLY listed. BiWaze Clear & Volara: Medicare is 'Considered Medically Necessary when criteria in applicable policy statements are met.' E0469/A7021 codes added to policy 07/17/2025.",
        "source": "https://wa-provider.kaiserpermanente.org/static/pdf/hosting/clinical/criteria/pdf/hfcwo.pdf"
    },
    # Kaiser WA Non-Medicare - E0469 Not Medically Necessary
    {
        "name": "Kaiser Permanente WA (Non-Medicare)",
        "type": "Commercial HMO",
        "coverage": "Case Review - Prior Auth Needed",
        "prior_auth": "Yes",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "07/17/2025",
        "policy_num": "HFCWO Criteria",
        "notes": "E0469 EXPLICITLY listed. For Non-Medicare plans: BiWaze Clear & Volara are 'Not medically Necessary - experimental, investigational or unproven.' E0469/A7021 codes added to policy 07/17/2025. May be reviewed on case-by-case basis.",
        "source": "https://wa-provider.kaiserpermanente.org/static/pdf/hosting/clinical/criteria/pdf/hfcwo.pdf"
    },
    # UnitedHealthcare Commercial - E0469 OLE UNPROVEN
    {
        "name": "UnitedHealthcare (Commercial)",
        "type": "Commercial",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (OLE devices)",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. OLE combination devices (CPEP, CHFO, nebulization) considered UNPROVEN and NOT MEDICALLY NECESSARY effective 03/01/2025. HFCWO covered for CF, bronchiectasis, neuromuscular.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/commercial-policies/commercial-medical-drug-policies.html"
    },
    # UnitedHealthcare Community Plan - E0469
    {
        "name": "UnitedHealthcare Community Plan (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "Airway Clearance Devices CS",
        "notes": "E0469 EXPLICITLY listed. HFCWO unproven and not medically necessary for conditions other than CF, bronchiectasis, neuromuscular disease.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/airway-clearance-devices-cs.pdf"
    },
    # UnitedHealthcare Individual Exchange - E0469
    {
        "name": "UnitedHealthcare Individual Exchange",
        "type": "ACA Exchange",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (OLE devices)",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed in applicable codes: A7021, A7025, A7026, E0469, E0481, E0483. HFCWO and IPV devices policy.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/exchange-policies/exchanges-medical-drug-policies.html"
    },
    # UMR (UHC TPA) - E0469
    {
        "name": "UMR (UnitedHealthcare TPA)",
        "type": "Commercial TPA",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (OLE devices)",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed in applicable codes. Same policy as UHC Commercial - OLE devices unproven, HFCWO covered for specific conditions.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/commercial-policies/umr-medical-drug-policies.html"
    },
    # Surest (UHC) - E0469
    {
        "name": "Surest (UnitedHealthcare)",
        "type": "Commercial",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (OLE devices)",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. Same UHC policy applies - OLE devices considered unproven.",
        "source": "https://ams-gateway.uhcprovider.com/en/policies-protocols/commercial-policies/surest-medical-drug-policies.html"
    },
    # UnitedHealthcare Oxford - E0469
    {
        "name": "UnitedHealthcare Oxford",
        "type": "Commercial",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices OHP",
        "notes": "E0469 EXPLICITLY listed. 2-month rental trial for HFCWO. Coverage for CF, bronchiectasis, neuromuscular disease. OLE devices unproven.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/oxford/airway-clearance-devices-ohp.pdf"
    },
    # UHC New Jersey Medicaid - E0469
    {
        "name": "UnitedHealthcare New Jersey (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Not Specified",
        "date": "11/01/2025",
        "policy_num": "NJ-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Policy effective 11/01/2025. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/new-jersey-health-plans/nj-comm-plan-home/nj-cp-policies/medicaid-community-state-policies-nj.html"
    },
    # UHC Louisiana Medicaid - E0469
    {
        "name": "UnitedHealthcare Louisiana (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - Limited Conditions",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "LA-CS Policy",
        "notes": "E0469 EXPLICITLY listed. HFCWO unproven and not medically necessary for conditions other than CF, bronchiectasis, neuromuscular disease. Uses InterQual criteria.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/la/airway-clearance-devices-la-cs.pdf"
    },
    # Rocky Mountain Health Plans - E0469
    {
        "name": "Rocky Mountain Health Plans (Colorado)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "03/01/2025",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. Policy addresses HFCWO, IPV devices. Codes A7021, A7025, A7026, E0469, E0481, E0483. Partners with UHC Community Plan.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/exchange-policies/rmhp-medical-policies.html"
    },
    # Cigna - E0469 COVERED
    {
        "name": "Cigna",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Policy 0069",
        "notes": "E0469 EXPLICITLY included in Airway Clearance Devices policy. Covered for cystic fibrosis and chronic bronchiectasis (>6 months, CT confirmed). Must have failed conventional therapies.",
        "source": "https://static.cigna.com/assets/chcp/pdf/coveragePolicies/medical/mm_0069_coveragepositioncriteria_airway_clearance_devices.pdf"
    },
    # Humana Medicare Advantage - E0469 Prior Auth Required
    {
        "name": "Humana Medicare Advantage",
        "type": "Medicare Advantage",
        "coverage": "Covered with Prior Auth",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "MA Prior Auth List",
        "notes": "E0469 EXPLICITLY requires prior authorization on 2026 Prior Auth list. 90-day grace period for new enrollees in active treatment. Reference CP2023011.",
        "source": "https://assets.humana.com/is/content/humana/FINAL_Medicare%20and%20DSNP%20Prior%20Authorization%20and%20Notification%20List%20-%201-1-2026pdf"
    },
    # Humana Commercial - E0469 Investigational
    {
        "name": "Humana (Commercial)",
        "type": "Commercial",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 devices (BiWaze Clear, IPV, Volara) 'considered experimental/investigational - not identified as widely used and generally accepted for proposed uses.'",
        "source": "https://genhealth.ai/policy/humana/8eaae7e6-airway-clearance-devices"
    },
    # Anthem Connecticut - E0469
    {
        "name": "Anthem Blue Cross Connecticut",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CG-DME-43",
        "notes": "E0469 EXPLICITLY listed in High Frequency Chest Compression Devices for Airway Clearance policy (CG-DME-43). Prior authorization required.",
        "source": "https://genhealth.ai/policy/anthem-bluecross-ct/6001b967-cg-dme-43-high-frequency-chest-compression-devices-for-airway-clearance"
    },
    # Geisinger - E0469 COVERED
    {
        "name": "Geisinger Health Plan",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (some devices)",
        "not_med_necessary": "Yes (some devices)",
        "date": "Current",
        "policy_num": "MP045",
        "notes": "E0469 EXPLICITLY listed. Covers E0469, E0480, E0481, E0482. For neuromuscular disease with respiratory weakness. Pre-certification required. Some devices NOT covered - insufficient evidence.",
        "source": "https://www.geisinger.org/-/media/OneGeisinger/Files/Policy%20PDFs/MP/1-50/MP045%20High%20Frequency%20Chest%20Percussion%20Vest.ashx"
    },
    # Univera Healthcare NY - E0469 COVERED
    {
        "name": "Univera Healthcare (New York)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (replacements)",
        "date": "Current",
        "policy_num": "Policy 1.01.15",
        "notes": "E0469 EXPLICITLY listed. Also E0481, E0482, E0483 covered. Referral must be from pulmonologist. Contract dependent. Replacement of functioning device NOT covered.",
        "source": "https://www.univerahealthcare.com/documents/d/global/uni-prv-airway-clearance-devices-1"
    },
    # ABM BiWaze Reimbursement Reference - E0469
    {
        "name": "BiWaze Clear Reimbursement Guide (ABM)",
        "type": "Manufacturer Reference",
        "coverage": "Reference Only",
        "prior_auth": "Varies by Payer",
        "investigational": "Many Payers - Yes",
        "not_med_necessary": "Many Payers - Yes",
        "date": "10/01/2024",
        "policy_num": "E0469 + A7021",
        "notes": "E0469 effective 10/1/2024 for OLE therapy. A7021 for monthly disposables. Industry reference notes many payers consider OLE devices investigational.",
        "source": "https://resources.abmrc.com/biwaze-clear-reimbursement-hcpcs-codes-e0469-a7021-for-ole-therapy"
    },
    # Noridian JD DME - E0469 billing guidance
    {
        "name": "Noridian Medicare (JD DME)",
        "type": "Medicare MAC",
        "coverage": "Case-by-Case (No LCD)",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "01/07/2025",
        "policy_num": "DMD Article 2025",
        "notes": "E0469 billing article published for JD region (AZ, CO, HI, NV, NM, OK, OR, TX, WA, WY, Pacific territories). No LCD - claims reviewed individually.",
        "source": "https://med.noridianmedicare.com/web/jddme/policies/dmd-articles/2025/lung-expansion-airway-clearance-continuous-high-frequency-oscillation-and-nebulization-device-hcpcs-code-e0469-correct-coding-and-billing-of-hcpcs"
    },
    # UHC North Carolina Community Plan - E0469
    {
        "name": "UnitedHealthcare North Carolina (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "04/01/2025",
        "policy_num": "NC-CS Policy",
        "notes": "E0469 EXPLICITLY listed. North Carolina Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/north-carolina-health-plans/nc-comm-plan-home/nc-cp-policies/nc-medicaid-community-state-policies.html"
    },
    # UHC Pennsylvania Community Plan - E0469
    {
        "name": "UnitedHealthcare Pennsylvania (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "01/01/2026",
        "policy_num": "PA-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Pennsylvania Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/pennsylvania-health-plans/pa-comm-plan-home/pa-cp-policies/medicaid-community-state-policies-pa.html"
    },
    # UHC Tennessee Community Plan - E0469
    {
        "name": "UnitedHealthcare Tennessee (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "TN-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Tennessee Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/tennessee-health-plans/tn-comm-plan-home/tn-cp-policies/medicaid-community-state-policies-tn.html"
    },
    # Excellus BCBS - E0469 COVERED
    {
        "name": "Excellus BlueCross BlueShield (NY)",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (replacements)",
        "date": "Current",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed along with E0481, E0482, E0483. Referral must be from pulmonologist. Contract dependent. Replacement of functioning device NOT covered.",
        "source": "https://www.excellusbcbs.com/documents/d/global/exc-prv-airway-clearance-devices"
    },
    # Blue Cross NC - E0469 INVESTIGATIONAL
    {
        "name": "Blue Cross Blue Shield North Carolina",
        "type": "BCBS",
        "coverage": "Investigational",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "03/2025 (reviewed)",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "E0469 EXPLICITLY listed in Applicable Codes. Volara System OLE 3-in-1 combined therapy device 'is considered INVESTIGATIONAL.' Originated 03/1998, last review 03/2025.",
        "source": "https://www.bluecrossnc.com/providers/policies-guidelines-codes/commercial/home-health-dme/updates/oscillatory-devices-for-treatment-of-respiratory-conditions"
    },
    # BCBS Kansas - E0469 Listed
    {
        "name": "Blue Cross Blue Shield Kansas",
        "type": "BCBS",
        "coverage": "Partial - Some Investigational",
        "prior_auth": "Yes",
        "investigational": "Yes (some conditions)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "E0469 listed in Oscillatory Devices for CF and other respiratory disorders policy. Coverage varies by condition. COPD without bronchiectasis may be investigational.",
        "source": "https://www.bcbsks.com/medical-policies/oscillatory-devices-treatment-cystic-fibrosis-and-other-respiratory-disorders"
    },
    # BCBS Massachusetts - E0469 policy
    {
        "name": "Blue Cross Blue Shield Massachusetts",
        "type": "BCBS",
        "coverage": "Partial - Some Investigational",
        "prior_auth": "Yes",
        "investigational": "Yes (some uses)",
        "not_med_necessary": "Not Specified",
        "date": "01/2025",
        "policy_num": "Policy 120",
        "notes": "Oscillatory devices for CF and respiratory conditions. Medical policy updates announced effective January 2025. Some uses considered investigational.",
        "source": "https://www.bluecrossma.org/medical-policies/sites/g/files/csphws2091/files/acquiadam-assets/120%20Oscillatory%20Devices%20for%20the%20Treatment%20of%20Cystic%20Fibrosis%20and%20Other%20Respiratory%20Conditions%20prn.pdf"
    },
    # Health Plan of Nevada - E0469 Listed
    {
        "name": "Health Plan of Nevada",
        "type": "Regional HMO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. Policy addresses HFCWO and IPV devices. Applicable codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://healthplanofnevada.com/provider/medical-policies"
    },
    # Sierra Health and Life (Nevada) - E0469 Listed
    {
        "name": "Sierra Health and Life (Nevada)",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed (UHC affiliate). Same policy as Health Plan of Nevada. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.sierrahealthandlife.com/provider/medical-policies"
    },
    # UHC Kentucky Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Kentucky (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "KY-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Kentucky Community Plan policy. Uses InterQual criteria. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/kentucky-health-plans/ky-comm-plan-home/ky-cp-policies/medicaid-community-state-policies-ky.html"
    },
    # UHC Texas Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Texas (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "TX-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Texas Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/texas-health-plans/tx-comm-plan-home/tx-cp-policies/medicaid-community-state-policies-tx.html"
    },
    # UHC Arizona Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Arizona (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "AZ-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Arizona Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/arizona-health-plans/az-comm-plan-home/az-cp-policies/medicaid-community-state-policies-az.html"
    },
    # UHC Michigan Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Michigan (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "MI-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Michigan Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/michigan-health-plans/mi-comm-plan-home/mi-cp-policies/medicaid-community-state-policies-mi.html"
    },
    # UHC Ohio Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Ohio (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "OH-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Ohio Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/ohio-health-plans/oh-comm-plan-home/oh-cp-policies/medicaid-community-state-policies-oh.html"
    },
    # UHC Virginia Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Virginia (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "VA-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Virginia Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/virginia-health-plans/va-comm-plan-home/va-cp-policies/medicaid-community-state-policies-va.html"
    },
    # UHC Wisconsin Community Plan - E0469 Listed
    {
        "name": "UnitedHealthcare Wisconsin (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - OLE Unproven",
        "prior_auth": "Yes",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Yes (other conditions)",
        "date": "11/01/2025",
        "policy_num": "WI-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Wisconsin Community Plan policy. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/wisconsin-health-plans/wi-comm-plan-home/wi-cp-policies/medicaid-community-state-policies-wi.html"
    },
    # California Medi-Cal - E0469 Covered
    {
        "name": "California Medi-Cal (Medicaid)",
        "type": "Medicaid",
        "coverage": "Covered with Limits",
        "prior_auth": "Not Specified",
        "investigational": "No",
        "not_med_necessary": "No",
        "date": "Current",
        "policy_num": "DME Billing Codes Manual",
        "notes": "E0469 EXPLICITLY listed in Medi-Cal DME Billing Codes. Frequency limit: 1 in 5 years. Authorization limited to lowest cost item meeting patient needs per Title 22 CCR Section 51321(g).",
        "source": "https://mcweb.apps.prd.cammis.medi-cal.ca.gov/file/manual?fn=duracdfre.pdf"
    },
    # Humana Medicaid - E0469 Investigational
    {
        "name": "Humana Medicaid",
        "type": "Medicaid MCO",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Multi-Function OLE Therapy",
        "notes": "E0469 EXPLICITLY mentioned. Multi-function oscillation lung expansion therapy (BiWaze Clear, Volara) considered investigational/experimental. Absence of treatment guidelines and clinical literature.",
        "source": "https://assets.humana.com/is/content/humana/Medicaid_Multi-Function_Oscillation_Lung_Expansion_Therapypdf"
    },
    # Aetna - E0469 INVESTIGATIONAL
    {
        "name": "Aetna",
        "type": "Commercial",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CPB 0067",
        "notes": "E0469 referenced in Chest Physiotherapy and Airway Clearance Devices policy. Volara System OLE therapy device considered investigational/experimental due to insufficient evidence supporting effectiveness.",
        "source": "https://www.aetna.com/cpb/medical/data/1_99/0067.html"
    },
    # EmblemHealth - E0469 Rental Only
    {
        "name": "EmblemHealth (New York)",
        "type": "Regional",
        "coverage": "Covered - Rental Only",
        "prior_auth": "May be Required",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "08/20/2024",
        "policy_num": "DME Rental vs Purchase Policy",
        "notes": "E0469 EXPLICITLY added to Items Eligible for Rental Only table effective 8/20/2024. 13-month capped rental period. Preauthorization may be required depending on network and financial risk entity.",
        "source": "https://www.emblemhealth.com/content/dam/emblemhealth/pdfs/provider/reimbursement-policies/dme-rental-purchase-emblemhealth.pdf"
    },
    # Moda Health (Oregon/Alaska) - E0469 Listed
    {
        "name": "Moda Health (Oregon/Alaska)",
        "type": "Regional",
        "coverage": "Covered with Prior Auth",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "11/03/2025",
        "policy_num": "MHMNC - HFCWO Devices",
        "notes": "E0469 EXPLICITLY listed under Airway Clearance Devices / Chest Percussors / Vest / IPV. Prior authorization required. Uses Moda Health Medical Necessity Criteria (MHMNC) for High Frequency Chest Wall Oscillation Devices.",
        "source": "https://www.modahealth.com/-/media/modahealth/shared/provider/prior-authorization/pre-auth-list-commercial.pdf"
    },
    # BCBS Rhode Island - E0469 Listed
    {
        "name": "Blue Cross Blue Shield Rhode Island",
        "type": "BCBS",
        "coverage": "Case Review - Prior Auth Needed",
        "prior_auth": "Yes - Preauthorization recommended (Commercial), Required (MA)",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (Commercial)",
        "date": "10/2024",
        "policy_num": "HCPCS Level II Updates",
        "notes": "E0469 EXPLICITLY listed in October 2024 HCPCS updates. For Commercial: 'Not medically necessary' - preauthorization recommended. For Medicare Advantage: Subject to medical review, preauthorization required.",
        "source": "https://www.bcbsri.com/providers/update/additional-hcpcs-level-ii-code-changes-and-modifier-changes-4"
    },
    # Wellmark BCBS - E0469 OLE Investigational
    {
        "name": "Wellmark BCBS (Iowa/South Dakota)",
        "type": "BCBS",
        "coverage": "Investigational",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. OLE devices (Volara, BiWaze Clear) 'considered investigational because the evidence is insufficient to determine the technology results in an improvement in the net health outcomes.'",
        "source": "https://digital-assets.wellmark.com/adobe/assets/urn:aaid:aem:8aa2f47a-61ff-4beb-9f1b-219c24adb04e/original/as/Airway-Clearance-Devices.pdf"
    },
    # BCBS Michigan - E0469 Investigational (External Review Case)
    {
        "name": "Blue Cross Blue Shield of Michigan",
        "type": "BCBS",
        "coverage": "Investigational - Experimental",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "09/2024",
        "policy_num": "DIFS Case 227614",
        "notes": "Volara System (E0469) determined 'experimental/investigational' in Michigan DIFS external review case. IRO found 'not yet FDA approved or FDA certified and there are no high-quality medical studies available.' Denial upheld.",
        "source": "https://www.michigan.gov/difs/-/media/Project/Websites/difs/PRIRA/2024/September/BCBSM_227614.pdf"
    },
    # Lifewise (Washington) - E0469 in Fee Schedule
    {
        "name": "Lifewise (Washington)",
        "type": "Commercial",
        "coverage": "Covered - Fee Schedule",
        "prior_auth": "Not Specified",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "07/15/2025",
        "policy_num": "Fee Schedule Maintenance Update",
        "notes": "E0469 EXPLICITLY listed in fee schedule maintenance update. Fixed pricing aligned with CMS fee schedules. Modifiers: KR, NR, NU, RR, UE. Contract-specific compensation exhibits may apply.",
        "source": "https://providernews.lifewise.com/fee-schedule-maintenance-update/"
    },
    # Fidelis Care (New York Medicaid MCO) - E0469 in Authorization Grid
    {
        "name": "Fidelis Care (New York Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Covered with Prior Auth",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "DME Authorization Grid",
        "notes": "E0469 EXPLICITLY listed in DME Authorization Grid for Medicaid, Child Health Plus, Essential Plan, and QHP. Described as 'Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device'. Prior authorization required.",
        "source": "https://www.fideliscare.org/Provider/Provider-Resources/Authorization-Grids"
    },
    # Kentucky Medicaid MSEA - E0469 Referenced
    {
        "name": "Kentucky Medicaid MSEA",
        "type": "Medicaid",
        "coverage": "Covered - Per Fee Schedule",
        "prior_auth": "Check Fee Schedule",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "2025",
        "policy_num": "907 KAR 1:479",
        "notes": "E0469 EXPLICITLY referenced as applicable procedure code for airway clearance devices under MSEA (Medical Supplies, Equipment, and Appliances) regulation 907 KAR 1:479. 2025 MSEA Fee Schedule available.",
        "source": "https://www.chfs.ky.gov/agencies/dms/dpo/bpb/Pages/dme.aspx"
    },
    # Capital Blue Cross (Pennsylvania) - E0469 Experimental/Investigational
    {
        "name": "Capital Blue Cross (Pennsylvania)",
        "type": "BCBS",
        "coverage": "NOT COVERED - Experimental/Investigational",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "02/01/2026",
        "policy_num": "MP 4.002",
        "notes": "E0469 EXPLICITLY listed in Experimental and Investigational Procedures policy (MP 4.002) effective 2/1/2026. Service considered experimental/investigational when: not FDA approved, subject to investigational application, experts indicate further research needed, or not medically necessary except with investigational treatment.",
        "source": "https://www.capbluecross.com/wps/wcm/connect/prod_nws.capblue.com29556/12a40734-7069-4e43-8944-30b8c7dde992/medical-policy-experimental-and-investigational-procedures.pdf?MOD=AJPERES"
    },
    # BCBS Texas - E0469 EIU or Clinical Review
    {
        "name": "Blue Cross Blue Shield Texas",
        "type": "BCBS",
        "coverage": "Varies - EIU or Clinical Review",
        "prior_auth": "Recommended Clinical Review",
        "investigational": "Yes (some plans)",
        "not_med_necessary": "Plan Dependent",
        "date": "Current",
        "policy_num": "EIU Medical Policy / Fee Schedule",
        "notes": "E0469 EXPLICITLY listed. Coverage varies by plan: listed as Non-Reimbursable Experimental, Investigational and/or Unproven (EIU) for some plans, or requires Recommended Clinical Review for medical necessity determination. Fee schedule rate: $9,003.78 purchase (NU), $900.38 monthly rental.",
        "source": "https://www.bcbstx.com/provider/claims/claims-eligibility/fee-schedule"
    },
    # BCBS Illinois - E0469 EIU Non-Reimbursable
    {
        "name": "Blue Cross Blue Shield Illinois",
        "type": "BCBS",
        "coverage": "NOT COVERED - EIU Non-Reimbursable",
        "prior_auth": "Recommended Clinical Review",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "EIU Medical Policy",
        "notes": "E0469 EXPLICITLY listed under Non-Reimbursable/Experimental, Investigational, Unproven (EIU) services. Not typically covered as standard approved medical device. Subject to strict review. May require Recommended Clinical Review (formerly predetermination) if submitted.",
        "source": "https://www.bcbsil.com/provider/claims/claims-eligibility/fee-schedule"
    },
    # BCBS New Mexico - E0469 Recommended Clinical Review
    {
        "name": "Blue Cross Blue Shield New Mexico",
        "type": "BCBS",
        "coverage": "Prior Auth Required - Clinical Review",
        "prior_auth": "Yes - Recommended Clinical Review",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Predetermination List",
        "notes": "E0469 EXPLICITLY included in BCBSNM's Recommended Clinical Review (Predetermination) list. Requires prior authorization to determine medical necessity before coverage decision.",
        "source": "https://www.bcbsnm.com/provider/education-reference/education/tools/fee-schedules-availity"
    },
    # BCBS Nebraska - E0469 Medicare Advantage Prior Auth
    {
        "name": "Blue Cross Blue Shield Nebraska",
        "type": "BCBS",
        "coverage": "Prior Auth Required (MA)",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "MA-X-077",
        "notes": "E0469 EXPLICITLY listed in Medicare Advantage Medical Prior Authorization policy (MA-X-077). Prior authorization required for coverage determination.",
        "source": "https://www.nebraskablue.com/Providers/Policies-and-Procedures/Medicare-Advantage/MA-Medical-Prior-Authorization"
    },
    # BCBS Minnesota - E0469 Prior Auth via Evicore
    {
        "name": "Blue Cross Blue Shield Minnesota",
        "type": "BCBS",
        "coverage": "Prior Auth Required",
        "prior_auth": "Yes - via Evicore",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2024",
        "policy_num": "DME Code List",
        "notes": "E0469 and A7021 EXPLICITLY listed as requiring prior authorization. PA managed through Evicore (Evernorth). Code list published 11/15/2024, effective 01/01/2024.",
        "source": "https://www.evicore.com/sites/default/files/resources/2024-11/BCBSMN_DME_CodeList_Eff01.01.2024_Pub11.15.2024.pdf"
    },
    # BCBS Vermont - E0469 Investigational
    {
        "name": "Blue Cross Blue Shield Vermont",
        "type": "BCBS",
        "coverage": "Investigational",
        "prior_auth": "N/A (investigational)",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2025",
        "policy_num": "10.01.VT204",
        "notes": "E0469 EXPLICITLY listed as 'Investigational' in Medical Policy 10.01.VT204 coding table. Effective 01/01/2025.",
        "source": "https://www.bluecrossvt.org/sites/default/files/2024-12/Investigational%20Services_Publication_01.01.2025AM.pdf"
    },
    # CareSource Ohio Medicaid - E0469 Prior Auth Required
    {
        "name": "CareSource Ohio (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Prior Auth Required",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "04/01/2025",
        "policy_num": "P-3504381",
        "notes": "E0469 and A7021 EXPLICITLY listed as requiring prior authorization effective 04/01/2025. New CMS codes added to PA list per provider notice dated 01/31/2025.",
        "source": "https://www.caresource.com/documents/oh-med-p-3504381-2025-prior-authorization-requirement-updates.pdf"
    },
    # Select Health Utah (Medicare) - E0469 Prior Auth Required
    {
        "name": "Select Health Utah (Medicare)",
        "type": "Medicare Advantage",
        "coverage": "Prior Auth Required",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "12/22/2025",
        "policy_num": "Medicare UT/ID Code List",
        "notes": "E0469 'Lung expansion airway clearance' EXPLICITLY listed as requiring prior authorization for Medicare members in Utah. Effective 12/22/2025.",
        "source": "https://selecthealth.org/content/dam/selecthealth/Provider/PDFs/preauthorization/non-covered-code-lists/medicare-ut-id.pdf"
    },
    # Select Health Idaho (Medicare) - E0469 Prior Auth Required
    {
        "name": "Select Health Idaho (Medicare)",
        "type": "Medicare Advantage",
        "coverage": "Prior Auth Required",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "12/22/2025",
        "policy_num": "Medicare UT/ID Code List",
        "notes": "E0469 'Lung expansion airway clearance' EXPLICITLY listed as requiring prior authorization for Medicare members in Idaho. Effective 12/22/2025.",
        "source": "https://selecthealth.org/content/dam/selecthealth/Provider/PDFs/preauthorization/non-covered-code-lists/medicare-ut-id.pdf"
    },
]

# Write data
for row_num, payer in enumerate(payer_data, 2):
    ws.cell(row=row_num, column=1, value=payer["name"])
    ws.cell(row=row_num, column=2, value=payer["type"])

    coverage_cell = ws.cell(row=row_num, column=3, value=payer["coverage"])
    if "Investigational" in payer["coverage"] or "NOT COVERED" in payer["coverage"]:
        coverage_cell.fill = not_covered_fill
    elif "Case-by-Case" in payer["coverage"]:
        coverage_cell.fill = case_by_case_fill
    elif "Partial" in payer["coverage"]:
        coverage_cell.fill = investigational_fill
    elif "Covered" in payer["coverage"]:
        coverage_cell.fill = covered_fill

    ws.cell(row=row_num, column=4, value=payer["prior_auth"])

    inv_cell = ws.cell(row=row_num, column=5, value=payer["investigational"])
    if payer["investigational"] == "Yes" or payer["investigational"].startswith("Yes"):
        inv_cell.fill = investigational_fill

    nmn_cell = ws.cell(row=row_num, column=6, value=payer["not_med_necessary"])
    if payer["not_med_necessary"] == "Yes" or payer["not_med_necessary"].startswith("Yes") or "Yes" in payer["not_med_necessary"]:
        nmn_cell.fill = not_covered_fill

    ws.cell(row=row_num, column=7, value=payer["date"])
    ws.cell(row=row_num, column=8, value=payer["policy_num"])
    ws.cell(row=row_num, column=9, value=payer["notes"])
    ws.cell(row=row_num, column=10, value=payer["source"])

    # Apply borders and alignment
    for col in range(1, 11):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Set column widths
column_widths = [38, 18, 28, 15, 25, 25, 18, 25, 70, 60]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Set row height for header
ws.row_dimensions[1].height = 30

# Freeze top row
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet(title="Summary")
ws2.cell(row=1, column=1, value="E0469 EXPLICIT Payer Policy Analysis")
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

# Count categories
total_payers = len([p for p in payer_data if p["type"] != "Manufacturer Reference"])
not_covered = sum(1 for p in payer_data if "NOT COVERED" in p["coverage"])
covered = sum(1 for p in payer_data if "Covered" in p["coverage"] and "NOT COVERED" not in p["coverage"])
case_by_case = sum(1 for p in payer_data if "Case-by-Case" in p["coverage"])
partial = sum(1 for p in payer_data if "Partial" in p["coverage"])
investigational_yes = sum(1 for p in payer_data if p["investigational"].startswith("Yes") or p["investigational"] == "Yes")

summary_data = [
    ("", ""),
    ("HCPCS Code:", "E0469"),
    ("Description:", "Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device"),
    ("Effective Date:", "October 1, 2024 (CMS introduced code)"),
    ("Related Code:", "A7021 (monthly disposables for E0469)"),
    ("Devices:", "Volara, BiWaze Clear, MetaNeb (OLE therapy devices)"),
    ("", ""),
    ("IMPORTANT NOTE:", "This spreadsheet contains ONLY payers that EXPLICITLY mention E0469 in their published policies."),
    ("", ""),
    ("COVERAGE SUMMARY:", ""),
    ("Total Payers with Explicit E0469 Policies:", str(total_payers)),
    ("NOT COVERED (Investigational/Experimental):", str(not_covered)),
    ("Covered with Criteria:", str(covered)),
    ("Partial Coverage (OLE unproven, HFCWO may be covered):", str(partial)),
    ("Case-by-Case Review (No LCD):", str(case_by_case)),
    ("Marked as Investigational:", str(investigational_yes)),
    ("", ""),
    ("KEY FINDINGS:", ""),
    ("1.", "Medicare (CMS) has NO LCD or NCD for E0469 - claims reviewed case-by-case"),
    ("2.", "Most payers that explicitly mention E0469 classify OLE devices as INVESTIGATIONAL"),
    ("3.", "UnitedHealthcare (all product lines) explicitly lists E0469 but considers OLE devices 'unproven'"),
    ("4.", "Regional plans (Medica, HealthPartners, Kaiser WA Non-Medicare) explicitly do NOT cover E0469"),
    ("5.", "BCBS Florida and Premera explicitly list E0469 as investigational/experimental"),
    ("6.", "Cigna Policy 0069 explicitly COVERS E0469 for CF and bronchiectasis"),
    ("7.", "Humana Medicare Advantage requires prior auth; Humana Commercial considers it investigational"),
    ("", ""),
    ("OLE DEVICE COVERAGE STATUS:", ""),
    ("Volara System:", "Largely considered investigational (recalled April 2022)"),
    ("BiWaze Clear:", "Largely considered investigational (FDA 510(k) cleared 2022)"),
    ("MetaNeb System:", "Largely considered investigational for home use"),
    ("", ""),
    ("COLOR KEY:", ""),
    ("Green:", "Covered with criteria"),
    ("Yellow:", "Partial coverage or investigational status noted"),
    ("Red:", "Not covered or investigational/experimental"),
    ("Blue:", "Case-by-case review (no LCD)"),
    ("", ""),
    ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("Data Source:", "Web search of published payer policies - January/February 2026"),
]

for row_num, (label, value) in enumerate(summary_data, 2):
    ws2.cell(row=row_num, column=1, value=label)
    if label.endswith(":") or label.startswith("-"):
        ws2.cell(row=row_num, column=1).font = Font(bold=True)
    ws2.cell(row=row_num, column=2, value=value)

ws2.column_dimensions['A'].width = 50
ws2.column_dimensions['B'].width = 85

# Add sheet for payers searched but no explicit E0469 found
ws3 = wb.create_sheet(title="Searched - No E0469 Found")
ws3.cell(row=1, column=1, value="Payers Searched - No Explicit E0469 Policy Found")
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
ws3.cell(row=2, column=1, value="These payers were searched but do not have published policies that explicitly mention E0469.")
ws3.cell(row=2, column=1).font = Font(italic=True)

# Headers for searched payers
searched_headers = ["Payer Name", "Payer Type", "Notes", "Date Searched"]
for col, header in enumerate(searched_headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Payers searched but no explicit E0469 found
searched_no_e0469 = [
    # BCBS Plans
    {"name": "BCBS Tennessee", "type": "BCBS", "notes": "Oscillatory devices policy covers E0483, not E0469"},
    {"name": "BCBS Arkansas", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "BCBS Idaho", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "BCBS Montana", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Horizon BCBS NJ", "type": "BCBS", "notes": "Medical policies searched - no E0469 found"},
    {"name": "Empire BCBS (NY)", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Independence Blue Cross (PA)", "type": "BCBS", "notes": "Medical policy portal searched - no E0469 found"},
    {"name": "CareFirst BCBS (MD/DC/VA)", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Blue Shield California", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Highmark BCBS (PA/WV/DE)", "type": "BCBS", "notes": "Policy covers E0483 but not E0469"},
    {"name": "Regence BCBS (OR/ID/WA/UT)", "type": "BCBS", "notes": "Medical policy portal searched - no E0469 found"},
    # Commercial/Regional
    {"name": "Oscar Health", "type": "Commercial", "notes": "No explicit E0469 policy found"},
    {"name": "Bright Health", "type": "Commercial", "notes": "Exited insurance business - no policy found"},
    {"name": "Clover Health", "type": "Medicare Advantage", "notes": "No explicit E0469 policy found"},
    {"name": "Devoted Health", "type": "Medicare Advantage", "notes": "Prior auth list checked - E0469 not explicitly listed"},
    {"name": "Alignment Healthcare", "type": "Medicare Advantage", "notes": "No explicit E0469 policy found"},
    {"name": "First Health", "type": "Commercial PPO", "notes": "No explicit E0469 policy found"},
    {"name": "ConnectiCare", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Sentara Health Plans", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Harvard Pilgrim / Point32Health", "type": "Regional", "notes": "Medical necessity guidelines searched - no E0469 found"},
    {"name": "Tufts Health Plan / Point32Health", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Medical Mutual of Ohio", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Priority Health (MI)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "SelectHealth (UT)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "UPMC Health Plan (PA)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "MVP Health Care (NY/VT)", "type": "Regional", "notes": "Payment policies searched - no E0469 found"},
    {"name": "Quartz Health (WI)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Fallon Health (MA)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Magellan Health", "type": "Specialty", "notes": "DME policies searched - no E0469 found"},
    # Medicaid MCOs
    {"name": "Molina Healthcare", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found in multiple states"},
    {"name": "Centene / WellCare", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found"},
    {"name": "Amerigroup / Elevance", "type": "Medicaid MCO", "notes": "CG-DME-43 covers E0483, not E0469"},
    {"name": "AmeriHealth Caritas", "type": "Medicaid MCO", "notes": "Prior auth lists checked - no E0469 found"},
    {"name": "Superior Health Plan (TX)", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found"},
    # State Medicaid FFS Programs
    {"name": "Pennsylvania DHS Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Ohio ODM Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Michigan MDHHS Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "North Carolina NCTracks", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Virginia DMAS Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "New Jersey FamilyCare", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Arizona AHCCCS", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Washington Apple Health", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Kentucky DMS Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Louisiana LDH Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Indiana FSSA Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Connecticut DSS/HUSKY", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Nevada DHCFP Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Maryland Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Wisconsin ForwardHealth", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "South Carolina SCDHHS", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Oregon OHP", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Illinois HFS Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Georgia DCH Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Texas HHSC Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    {"name": "Florida AHCA Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed"},
    # Federal Programs
    {"name": "Tricare", "type": "Federal", "notes": "No explicit E0469 policy found - may follow Medicare guidance"},
    {"name": "VA Health Benefits", "type": "Federal", "notes": "No explicit E0469 policy found"},
    # Workers' Compensation
    {"name": "Federal OWCP", "type": "Workers' Comp", "notes": "No explicit E0469 policy found"},
    {"name": "NY State Workers' Comp", "type": "Workers' Comp", "notes": "Fee schedule not web-indexed"},
    {"name": "CA State Workers' Comp", "type": "Workers' Comp", "notes": "Fee schedule not web-indexed"},
    {"name": "TX State Workers' Comp", "type": "Workers' Comp", "notes": "Fee schedule not web-indexed"},
    # Anthem/Elevance (specific note)
    {"name": "Anthem BCBS (Multi-state)", "type": "BCBS/Elevance", "notes": "CG-DME-43 policy covers E0483 but E0469 NOT listed in applicable codes"},
    # Additional commercial payers searched Feb 2026
    {"name": "Centene / WellCare / Ambetter", "type": "Commercial/Medicaid MCO", "notes": "No explicit E0469 policy found"},
    {"name": "LifeWise / Cambia Health", "type": "Regional (WA)", "notes": "No explicit E0469 policy found"},
    {"name": "Allina Health Plan", "type": "Regional (MN)", "notes": "No explicit E0469 policy found"},
    {"name": "Sanford Health Plan", "type": "Regional (ND/SD/MN)", "notes": "No explicit E0469 policy found"},
    {"name": "BCBS Alabama", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "BCBS Montana", "type": "BCBS", "notes": "EIU policy updates but E0469 not confirmed"},
    {"name": "ConnectiCare (CT)", "type": "Regional", "notes": "EIU policy PDF unreadable - E0469 not confirmed"},
    {"name": "AvMed (FL)", "type": "Regional HMO", "notes": "No explicit E0469 policy found"},
    {"name": "Florida Health Care Plans (FHCP)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Alignment Healthcare", "type": "Medicare Advantage", "notes": "No explicit E0469 policy found"},
    {"name": "CDPHP (NY)", "type": "Regional HMO", "notes": "No explicit E0469 policy found"},
    {"name": "Fallon Health (MA)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "GuideWell / Florida Blue", "type": "BCBS (FL)", "notes": "Parent of Florida Blue - see BCBSFL policy for E0469"},
    {"name": "Neighborhood Health Plan (MA)", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found"},
    # Additional payers searched Feb 2026 (continued)
    {"name": "BCBS South Carolina", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "HAP Health Alliance Plan (MI)", "type": "Regional HMO", "notes": "No explicit E0469 policy found"},
    {"name": "Blue Shield California", "type": "BCBS", "notes": "Oscillatory devices policy found but E0469 not explicitly listed"},
    {"name": "Anthem California", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found"},
    {"name": "BCBS Wyoming", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Taro Health", "type": "Commercial", "notes": "No explicit E0469 policy found"},
    {"name": "Friday Health", "type": "Commercial", "notes": "No explicit E0469 policy found"},
    {"name": "Bind Benefits (UHC)", "type": "Commercial", "notes": "No explicit E0469 policy found"},
    {"name": "Aflac", "type": "Supplemental", "notes": "Supplemental insurance - no DME policy found"},
    {"name": "Lincoln Financial", "type": "Supplemental", "notes": "Supplemental insurance - no DME policy found"},
    {"name": "MetLife", "type": "Supplemental", "notes": "Supplemental insurance - no DME policy found"},
    {"name": "Humana SC Medicaid", "type": "Medicaid MCO", "notes": "Multi-function OLE policy references E0469 but PDF unreadable"},
    # Additional payers from user's list - searched Feb 2026
    {"name": "BCBS Louisiana (Louisiana Blue)", "type": "BCBS", "notes": "No explicit E0469 policy found - check iLinkBlue provider portal"},
    {"name": "Humana Puerto Rico", "type": "Regional", "notes": "No explicit E0469 policy - national Humana policy considers OLE investigational"},
    {"name": "Indiana Medicaid IHCP", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed for E0469 - check downloadable fee schedule"},
    {"name": "BCBS North Dakota", "type": "BCBS", "notes": "No explicit E0469 policy found"},
    {"name": "Providence Health Plan (Oregon)", "type": "Regional", "notes": "No explicit E0469 policy found - check DME/Supply list"},
    {"name": "Hawaii Medical Service Association (HMSA)", "type": "Regional", "notes": "No explicit E0469 policy found"},
    {"name": "Texas Children's Health Plan", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found"},
    {"name": "Anthem Wisconsin", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found - check Medical Policy Tool"},
    {"name": "BCBS California (Anthem Blue Cross CA)", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found"},
    {"name": "CalViva Health (California)", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found - check Medi-Cal fee schedule"},
    {"name": "iCare Health Plan (Wisconsin)", "type": "Medicaid MCO", "notes": "No explicit E0469 policy found"},
    {"name": "Colorado Medicaid HCPF", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed for E0469 - check Health First Colorado fee schedule"},
    {"name": "Select Health (Colorado)", "type": "Regional", "notes": "No explicit E0469 policy found - check DME policies document CR-41"},
    {"name": "BCBS Nevada", "type": "BCBS", "notes": "No explicit E0469 policy found (separate from Health Plan of Nevada)"},
    {"name": "Anthem Virginia", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found - DME limit changes noted but not E0469 specific"},
    {"name": "Anthem Maine", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found"},
    {"name": "Anthem New York", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found"},
    {"name": "Anthem Georgia", "type": "BCBS/Elevance", "notes": "No explicit E0469 policy found - administers GA State Health Benefit Plan"},
    {"name": "Mississippi Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed for E0469 - check DME fee schedule"},
    {"name": "Montana Medicaid", "type": "State Medicaid FFS", "notes": "Fee schedule not web-indexed for E0469 - check Jan 2025 DME fee schedule"},
]

for row_num, payer in enumerate(searched_no_e0469, 5):
    ws3.cell(row=row_num, column=1, value=payer["name"])
    ws3.cell(row=row_num, column=2, value=payer["type"])
    ws3.cell(row=row_num, column=3, value=payer["notes"])
    ws3.cell(row=row_num, column=4, value="Feb 2026")
    for col in range(1, 5):
        cell = ws3.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Set column widths for searched payers sheet
ws3.column_dimensions['A'].width = 40
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 60
ws3.column_dimensions['D'].width = 15

# Freeze header row
ws3.freeze_panes = "A5"

# Save workbook
output_path = "/Users/leahnoaeill/Downloads/MyClaude/data_export/E0469_Explicit_Payer_Policies.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
print(f"Total payers with EXPLICIT E0469 policies: {total_payers}")
print(f"  - Not Covered/Investigational: {not_covered}")
print(f"  - Covered with Criteria: {covered}")
print(f"  - Partial Coverage: {partial}")
print(f"  - Case-by-Case: {case_by_case}")
