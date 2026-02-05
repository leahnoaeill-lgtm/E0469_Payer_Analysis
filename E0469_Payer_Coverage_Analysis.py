#!/usr/bin/env python3
"""
Generate E0469 Payer Coverage Analysis Spreadsheet
HCPCS E0469: Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "E0469 Payer Coverage"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
covered_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
not_covered_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
investigational_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
case_by_case_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = [
    "Payer Name",
    "Payer Type",
    "Coverage Status",
    "Prior Auth Required",
    "Investigational/Experimental",
    "Not Medically Necessary",
    "Policy/Effective Date",
    "Policy Number",
    "Notes/Details",
    "Source URL"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Payer data compiled from research
payer_data = [
    # Medicare/CMS
    {
        "name": "Medicare (CMS/DME MACs)",
        "type": "Government",
        "coverage": "Case-by-Case",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "Ongoing",
        "policy_num": "No LCD/NCD",
        "notes": "No specific LCD or NCD for E0469. Each claim reviewed individually. Must document medical necessity. Capped rental item billed with RR modifier.",
        "source": "https://med.noridianmedicare.com/web/jadme/policies/dmd-articles/2025/lung-expansion-airway-clearance-continuous-high-frequency-oscillation-and-nebulization-device-hcpcs-code-e0469"
    },
    # UnitedHealthcare
    {
        "name": "UnitedHealthcare",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices Policy",
        "notes": "Commercial policy effective 01/01/2026 addresses airway clearance devices including E0469, E0481, E0483. Coverage for cystic fibrosis, bronchiectasis, neuromuscular conditions.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/commercial-policies/commercial-medical-drug-policies.html"
    },
    # Aetna
    {
        "name": "Aetna",
        "type": "Commercial",
        "coverage": "Partial - Some Investigational",
        "prior_auth": "Yes",
        "investigational": "Yes (some devices)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CPB 0067",
        "notes": "BiWaze Clear System considered experimental/investigational/unproven. HFCWO covered for some conditions (CF, bronchiectasis). Mixed findings noted for ALS.",
        "source": "https://www.aetna.com/cpb/medical/data/1_99/0067.html"
    },
    # Cigna
    {
        "name": "Cigna",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Policy 0069",
        "notes": "HFCWC covered for cystic fibrosis and chronic bronchiectasis (>6 months confirmed by CT). Must have failed conventional therapies. Policy 0069 Airway Clearance Devices.",
        "source": "https://static.cigna.com/assets/chcp/pdf/coveragePolicies/medical/mm_0069_coveragepositioncriteria_airway_clearance_devices.pdf"
    },
    # Humana
    {
        "name": "Humana",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "03/01/2023",
        "policy_num": "HUM-0310-020",
        "notes": "Coverage begins with 3-month rental trial. Eligible conditions: bronchiectasis, CF, ALS, muscular dystrophy, myasthenia gravis, spinal cord injuries. Documentation required.",
        "source": "https://genhealth.ai/policy/humana/8eaae7e6-airway-clearance-devices"
    },
    # Kaiser Permanente
    {
        "name": "Kaiser Permanente",
        "type": "Commercial/HMO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Criteria",
        "notes": "Non-Medicare: CF diagnosis primary criteria. Also covers quadriplegia with specific requirements. References Medicare LCD L33785 for Medicare members.",
        "source": "https://wa-provider.kaiserpermanente.org/static/pdf/hosting/clinical/criteria/pdf/hfcwo.pdf"
    },
    # Wellmark BCBS
    {
        "name": "Wellmark BCBS (Iowa/SD)",
        "type": "BCBS",
        "coverage": "Partial - OLE Investigational",
        "prior_auth": "Not Specified",
        "investigational": "Yes (OLE devices)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Airway Clearance Devices",
        "notes": "Oscillation lung expansion (OLE) devices like Volara and BiWaze Clear considered INVESTIGATIONAL due to insufficient evidence. Standard HFCWO may be covered.",
        "source": "https://digital-assets.wellmark.com/adobe/assets/urn:aaid:aem:8aa2f47a-61ff-4beb-9f1b-219c24adb04e/original/as/Airway-Clearance-Devices.pdf"
    },
    # Highmark
    {
        "name": "Highmark BCBS",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (Volara)",
        "not_med_necessary": "Yes (some devices)",
        "date": "02/22/2023",
        "policy_num": "MP-1141",
        "notes": "HFCWO covered for bronchiectasis with >2 exacerbations/year. Volara system NOT covered. Must document failure of standard treatments.",
        "source": "https://www.highmarkhealthoptions.com/content/dam/digital-marketing/en/highmark/highmarkhealthoptions/providers/medical-payment-policies/medical-policies/hho-de-mp-1141-highfrequencychestwalloscillationdevices_02222023.pdf"
    },
    # BCBS Florida
    {
        "name": "Blue Cross Blue Shield Florida",
        "type": "BCBS",
        "coverage": "Partial - OLE Investigational",
        "prior_auth": "Not Specified",
        "investigational": "Yes (OLE/CHFO)",
        "not_med_necessary": "Not Specified",
        "date": "09/25/2025 (reviewed)",
        "policy_num": "09-E0000-28",
        "notes": "OLE therapy and CHFO therapy in home setting lack evidence. Limited clinical trials, small sample sizes cited. MetaNeb/BiWaze devices not proven effective.",
        "source": "https://mcgs.bcbsfl.com/MCG?mcgId=09-E0000-28&pv=false"
    },
    # BCBS Tennessee
    {
        "name": "Blue Cross Blue Shield Tennessee",
        "type": "BCBS",
        "coverage": "Partial - Investigational for Some",
        "prior_auth": "Not Specified",
        "investigational": "Yes (COPD, other uses)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillating Devices Policy",
        "notes": "HFCWO and IPV devices considered INVESTIGATIONAL for COPD and conditions other than CF/bronchiectasis due to insufficient evidence on health outcomes.",
        "source": "https://www.bcbst.com/mpmanual/Oscillating_Devices_for_the_Treatment_of_Respiratory_Conditions.htm"
    },
    # BCBS North Carolina
    {
        "name": "Blue Cross NC",
        "type": "BCBS",
        "coverage": "OLE Investigational",
        "prior_auth": "Not Specified",
        "investigational": "Yes (Volara)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "Volara System OLE 3-in-1 combined therapy device is considered INVESTIGATIONAL. Standard HFCWO devices may have different coverage.",
        "source": "https://www.bluecrossnc.com/providers/policies-guidelines-codes/commercial/home-health-dme/updates/oscillatory-devices-for-treatment-of-respiratory-conditions"
    },
    # Premera Blue Cross
    {
        "name": "Premera Blue Cross",
        "type": "BCBS",
        "coverage": "OLE Investigational",
        "prior_auth": "Not Specified",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "10/01/2024 (codes added)",
        "policy_num": "1.01.539",
        "notes": "Oscillation and lung expansion devices (Volara, MetaNeb 4, BiWaze Clear) considered INVESTIGATIONAL. Added E0469 and A7021 codes effective 10/1/2024.",
        "source": "https://www.premera.com/medicalpolicies-individual/1.01.539.pdf"
    },
    # HealthPartners
    {
        "name": "HealthPartners",
        "type": "Regional",
        "coverage": "Investigational",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Airway Clearance Policy",
        "notes": "OLE therapy devices (E0469, A7021) considered investigational/experimental for home use. Prior auth not applicable as not covered.",
        "source": "https://www.healthpartners.com/ucm/groups/public/@hp/@public/@cc/documents/documents/aentry_045636.pdf"
    },
    # Centene (various subsidiaries)
    {
        "name": "Centene (Ambetter, WellCare, etc.)",
        "type": "Medicaid/Commercial",
        "coverage": "Per MCG Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Varies by State",
        "policy_num": "CP.MP.107",
        "notes": "Uses MCG criteria for DME. Covered if meets medical necessity per general DME policy. State-specific rules apply for Medicaid plans.",
        "source": "https://www.healthnet.com/content/dam/centene/policies/clinical-policies/CP.MP.107.pdf"
    },
    # Molina Healthcare
    {
        "name": "Molina Healthcare",
        "type": "Medicaid/Marketplace",
        "coverage": "Per Plan Document",
        "prior_auth": "Yes (generally)",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Varies by State",
        "policy_num": "N/A",
        "notes": "No specific E0469 policy found. Coverage determined by benefit document. Federal/state guidelines supersede internal policies.",
        "source": "https://www.molinahealthcare.com"
    },
    # California Medi-Cal
    {
        "name": "California Medi-Cal",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Dura Med Equip Manual",
        "notes": "May be medically necessary for: CF, bronchiectasis (confirmed by CT with daily productive cough >6 months or >2 exacerbations/year), neuromuscular disease.",
        "source": "https://mcweb.apps.prd.cammis.medi-cal.ca.gov/file/manual?fn=duraoxy.pdf"
    },
    # Minnesota Medicaid
    {
        "name": "Minnesota Medicaid (DHS)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "DHS16_152808",
        "notes": "Airway clearance devices for self-administered clearance for respiratory or neuromuscular conditions causing excessive mucus or difficulty clearing secretions.",
        "source": "https://www.dhs.state.mn.us/main/idcplg?IdcService=GET_DYNAMIC_CONVERSION&dDocName=DHS16_152808"
    },
    # Anthem
    {
        "name": "Anthem BCBS",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Policy",
        "notes": "Uses MCG Care Guidelines for UM decisions. Specific E0469 policy requires provider portal access. Prior auth typically required for DME.",
        "source": "https://www.anthem.com/provider/policies/clinical-guidelines/"
    },
    # Excellus BCBS (New York)
    {
        "name": "Excellus BCBS (New York)",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (replacement of functioning device)",
        "date": "Current",
        "policy_num": "Airway Clearance Devices",
        "notes": "Coverage contract dependent. Referral must be from pulmonologist. Covers CNY, Southern Tier, Rochester, Utica regions. Repair/replacement covered if compliant with use.",
        "source": "https://www.excellusbcbs.com/documents/d/global/exc-prv-airway-clearance-devices"
    },
    # Medical Mutual Ohio
    {
        "name": "Medical Mutual (Ohio)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "10/01/2022 (LCD ref)",
        "policy_num": "Policy #200508",
        "notes": "HFCWO System and IPV System policy. References LCD L33785 from CGS Administrators. Coverage for CF, bronchiectasis, neuromuscular conditions.",
        "source": "https://www.medmutual.com/-/media/MedMutual/Files/Providers/CorporateMedicalPolicies/200508_High-Frequency-Chest-Wall-Oscillation-System-and-Intrapulmonary-Percussive-Ventilation-System.pdf"
    },
    # HCSC (5 states)
    {
        "name": "HCSC (BCBS IL, TX, MT, NM, OK)",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Health Care Service Corporation operates BCBS in 5 states. Technology determinations based on peer-reviewed literature. Access policy at medicalpolicy.hcsc.com",
        "source": "https://medicalpolicy.hcsc.com/home.html?corpEntCd=HCSC"
    },
    # Regence BCBS
    {
        "name": "Regence BCBS (OR, WA, ID, UT)",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "References Oregon HERC guidance. HFCWO covered for CF (weak recommendation) and non-CF bronchiectasis. Must document failed chest PT/PEP therapy.",
        "source": "https://www.regence.com/provider/library/policies-guidelines/medical-policy"
    },
    # BCBS Kansas
    {
        "name": "Blue Cross Blue Shield Kansas",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (some conditions)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "Detailed policy for oscillatory devices for CF and other respiratory disorders. Coverage varies by condition. COPD without bronchiectasis may be investigational.",
        "source": "https://www.bcbsks.com/medical-policies/oscillatory-devices-treatment-cystic-fibrosis-and-other-respiratory-disorders"
    },
    # Blue Cross Massachusetts
    {
        "name": "Blue Cross Blue Shield Massachusetts",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (some uses)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Policy 120",
        "notes": "Oscillatory devices for CF and other respiratory conditions. Coverage criteria similar to Medicare LCD. Some uses may be investigational.",
        "source": "https://www.bluecrossma.org/medical-policies/sites/g/files/csphws2091/files/acquiadam-assets/120%20Oscillatory%20Devices%20for%20the%20Treatment%20of%20Cystic%20Fibrosis%20and%20Other%20Respiratory%20Conditions%20prn.pdf"
    },
    # Capital Blue Cross (PA)
    {
        "name": "Capital Blue Cross (Pennsylvania)",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "MIE Device Policy",
        "notes": "Mechanical insufflation-exsufflation device policy. Coverage for patients with impaired cough/secretion clearance due to neuromuscular conditions.",
        "source": "https://www.capbluecross.com/wps/wcm/connect/prod_nws.capblue.com29556/cc8cd29b-3a2f-4b25-a081-7d47ad23ff15/medical-policy-mechanical-insufflation-exsufflation-device.pdf"
    },
    # Moda Health (Oregon)
    {
        "name": "Moda Health (Oregon)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Criteria",
        "notes": "Medical necessity criteria for high frequency chest wall oscillation devices. Follows similar criteria to Medicare LCD guidelines.",
        "source": "https://www.modahealth.com/-/media/modahealth/shared/medical-necessity-criteria/HighFrequencyChestWallOscillationDevices.pdf"
    },
    # Iowa Medicaid
    {
        "name": "Iowa Medicaid (DHS)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "DME-009",
        "notes": "Clinical criteria for high frequency chest wall oscillation. Requires prior authorization. Coverage for CF, bronchiectasis, neuromuscular disease.",
        "source": "https://hhs.iowa.gov/media/315/download?inline="
    },
    # Oregon Health Authority
    {
        "name": "Oregon Health Plan (OHP)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "08/11/2022",
        "policy_num": "HERC Guidance",
        "notes": "Oregon HERC recommends coverage for CF (weak) and non-CF bronchiectasis. Must document failed chest PT/PEP or unavailable/not tolerated.",
        "source": "https://www.oregon.gov/oha/HPA/DSI-HERC/EvidenceBasedReports/High-frequency-chest-wall-oscillation-devices-Coverage-Guidance-APPROVED_8-11-22.pdf"
    },
    # Washington Health Care Authority
    {
        "name": "Washington Medicaid (HCA)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Policy",
        "notes": "Covers HFCWO for excessive mucus/difficulty clearing secretions. 12-month rental then purchase. Max rental $1,224.07/month (E0483). 1 device per lifetime.",
        "source": "https://www.hca.wa.gov"
    },
    # BCBS Federal Employee Program
    {
        "name": "BCBS Federal Employee Program (FEP)",
        "type": "Federal",
        "coverage": "Per FEP Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "FEP Medical Policy",
        "notes": "Federal Employee Program has separate medical policies. Coverage may differ from local BCBS plans. Check FEP policy portal.",
        "source": "https://www.fepblue.org/legal/policies-guidelines"
    },
    # CareSource (Ohio Medicaid)
    {
        "name": "CareSource (Ohio Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Per State Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "11/01/2025",
        "policy_num": "MM-1578",
        "notes": "Ohio Medicaid managed care. Follows state Medicaid policy for DME coverage. Prior authorization required for airway clearance devices.",
        "source": "https://www.caresource.com/documents/medicaid-oh-policy-medical-mm-1578-20251101"
    },
    # Massachusetts MassHealth
    {
        "name": "Massachusetts MassHealth",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Guidelines",
        "notes": "Guidelines for medical necessity determination for HFCWO air pulse generator system vest. Requires documented failure of standard therapies.",
        "source": "https://www.mass.gov/doc/guidelines-for-medical-necessity-determination-for-high-frequency-chest-wall-oscillation-air-pulse-generator-system-vest-0/download"
    },
    # BCBS North Dakota
    {
        "name": "Blue Cross Blue Shield North Dakota",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Policy",
        "notes": "High frequency chest wall oscillation devices policy. Coverage for CF, bronchiectasis, neuromuscular conditions meeting specific criteria.",
        "source": "https://www.bcbsnd.com/providers/policies-precertification/medical-policy/h/high-frequency-chest-wall-oscillation-devices"
    },
    # UnitedHealthcare Community Plan (Medicaid)
    {
        "name": "UnitedHealthcare Community Plan (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - Some Not Med Necessary",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (other conditions)",
        "date": "Current",
        "policy_num": "Airway Clearance Devices CS",
        "notes": "HFCWO unproven and not medically necessary for conditions other than CF, bronchiectasis, neuromuscular disease due to insufficient evidence.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/airway-clearance-devices-cs.pdf"
    },
    # UnitedHealthcare Oxford
    {
        "name": "UnitedHealthcare Oxford",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (other conditions)",
        "date": "Current",
        "policy_num": "Airway Clearance Devices OHP",
        "notes": "Oxford clinical policy for airway clearance devices. 2-month rental trial for HFCWO. Coverage for CF, bronchiectasis, neuromuscular disease.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/oxford/airway-clearance-devices-ohp.pdf"
    },
    # North Dakota Medicaid
    {
        "name": "North Dakota Medicaid",
        "type": "Medicaid",
        "coverage": "Per State Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CPAP Policy",
        "notes": "State Medicaid DME coverage policy. Check specific criteria for airway clearance devices including HFCWO and OLE devices.",
        "source": "https://www.hhs.nd.gov/sites/www/files/documents/DME/policy-cpap.pdf"
    },
    # TRICARE
    {
        "name": "TRICARE (Military)",
        "type": "Federal/Military",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "TRICARE Policy Manual",
        "notes": "HFCWO devices covered for CF, bronchiectasis, neuromuscular disease with prescription. Check TRICARE manuals at manuals.health.mil for specific criteria.",
        "source": "https://www.tricare.mil/CoveredServices"
    },
    # Veterans Affairs
    {
        "name": "Veterans Affairs (VA)",
        "type": "Federal/VA",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "VA Prosthetics",
        "notes": "Veterans may qualify through VA health benefits. Contact local VA facility prosthetics department. Generally follows Medicare guidelines.",
        "source": "https://www.va.gov/"
    },
    # Louisiana Medicaid / Healthy Blue
    {
        "name": "Louisiana Medicaid (Healthy Blue)",
        "type": "Medicaid MCO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "02/11/2022",
        "policy_num": "CG-DME-43",
        "notes": "HFCWO for airway clearance. Device designed for self-therapy with air pulse delivery system and inflatable vest. Requires documented need.",
        "source": "https://ldh.la.gov/assets/medicaid/MCPP/2.11.22/1253_HBL_CG_DME_43_High_Frequency_Chest_Compression_Devices_for_Airway_Clearance.pdf"
    },
    # Amerigroup (Multi-state)
    {
        "name": "Amerigroup (Multi-state)",
        "type": "Medicaid MCO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CG-DME-43",
        "notes": "HFCC devices covered when: FDA cleared, documented need for airway clearance, neuromuscular disorder affecting cough ability with history of pneumonia.",
        "source": "https://provider.amerigroup.com/dam/medpolicies/amerigroup/active/guidelines/gl_pw_d073857.html"
    },
    # BCBS South Carolina
    {
        "name": "Blue Cross Blue Shield South Carolina",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (insufficient evidence)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "HFCWO covered for CF, chronic diffuse bronchiectasis, neuromotor disorders when standard CPT failed/unavailable. Evidence deemed insufficient for other uses.",
        "source": "https://www.southcarolinablues.com/web/public/brands/medicalpolicyhb/external-policies/oscillatory-devices-for-the-treatment-of-cystic-fibrosis-and-other-respiratory-conditions/"
    },
    # Priority Health Michigan
    {
        "name": "Priority Health (Michigan)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Policy 91110-R19",
        "notes": "DME policy includes E0469. Coverage subject to member benefits. Must have FDA approval, be widely accepted, most appropriate level of care.",
        "source": "https://www.priorityhealth.com/-/media/priorityhealth/documents/medical-policies/91110.pdf"
    },
    # Presbyterian Health Plan New Mexico
    {
        "name": "Presbyterian Health Plan (New Mexico)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "MPM 4.3",
        "notes": "Respiratory Devices policy. Follows LCD L33800 for respiratory assist devices. Check Medical Policy Manual for specific criteria.",
        "source": "https://www.phs.org/providers/Documents/Durable-Medical-Equipment-Respiratory-MPM-4-3.pdf"
    },
    # Minnesota MHCP - E0469 NOT COVERED
    {
        "name": "Minnesota MHCP (E0469 specific)",
        "type": "Medicaid",
        "coverage": "NOT COVERED",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "DHS16_152808",
        "notes": "MHCP does NOT cover lung expansion airway clearance devices (E0469) for any indication - not standard in community care, substantive research lacking.",
        "source": "https://www.dhs.state.mn.us/main/idcplg?IdcService=GET_DYNAMIC_CONVERSION&dDocName=DHS16_152808"
    },
    # UHC Louisiana Community Plan
    {
        "name": "UnitedHealthcare Louisiana (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - Limited Conditions",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (other conditions)",
        "date": "Current",
        "policy_num": "LA-CS Policy",
        "notes": "HFCWO unproven and not medically necessary for conditions other than CF, bronchiectasis, neuromuscular disease. Uses InterQual criteria.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/la/airway-clearance-devices-la-cs.pdf"
    },
    # UHC Tennessee Community Plan
    {
        "name": "UnitedHealthcare Tennessee (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Partial - Limited Conditions",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (other conditions)",
        "date": "Current",
        "policy_num": "TN-CS Policy",
        "notes": "State-specific policy for Tennessee Medicaid. Coverage limited to specific qualifying conditions with documented need.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/tn/airway-clearance-devices-tn-cs.pdf"
    },
    # UHC New Mexico Community Plan
    {
        "name": "UnitedHealthcare New Mexico (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "NM-CS Policy",
        "notes": "New Mexico-specific Medicaid policy for airway clearance devices. Prior authorization required.",
        "source": "https://www.uhcprovider.com/content/dam/provider/docs/public/policies/medicaid-comm-plan/nm/airway-clearance-devices-nm-cs.pdf"
    },
    # North Carolina Medicaid
    {
        "name": "North Carolina Medicaid",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "05/18/2021 (updated)",
        "policy_num": "Policy 5A-2",
        "notes": "Respiratory Equipment and Supplies policy. Edit for HCPCS E0483 implemented. Check current policy for E0469 criteria.",
        "source": "https://medicaid.ncdhhs.gov/5a-2-respiratory-equipment-and-supplies/download"
    },
    # New York Medicaid eMedNY
    {
        "name": "New York Medicaid (eMedNY)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "DME Manual",
        "notes": "DME procedure codes and coverage guidelines. Prior authorization required for airway clearance devices.",
        "source": "https://www.emedny.org/providermanuals/dme/pdfs/dme_procedure_codes.pdf"
    },
    # ConnectiCare
    {
        "name": "ConnectiCare (Connecticut)",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Pre-authorization required for DME. Clinical evidence must show patient meets criteria. Check portal at connecticare.com/providers/our-policies/medical",
        "source": "https://www.connecticare.com/providers/our-policies/medical"
    },
    # Harvard Pilgrim
    {
        "name": "Harvard Pilgrim Health Care",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Medical Review Criteria",
        "notes": "Prior authorization based on clinical evidence. Contact 1-888-888-4742 for specific criteria. Medicare Advantage (Stride) has different requirements.",
        "source": "https://www.harvardpilgrim.org/public/prior-authorization-medical-review-criteria"
    },
    # Unicare
    {
        "name": "Unicare",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CG-DME-43",
        "notes": "High frequency chest compression devices for airway clearance. Similar criteria to Amerigroup/Anthem affiliated plans.",
        "source": "https://www.unicare.com/dam/medpolicies/unicare/active/guidelines/gl_pw_d073857.html"
    },
    # BCBS Alabama
    {
        "name": "Blue Cross Blue Shield Alabama",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Access medical policies at providers.bcbsal.org. Policy specifics require portal access.",
        "source": "https://providers.bcbsal.org/portal/resources/-/resources/category/319574"
    },
    # BCBS Arizona
    {
        "name": "Blue Cross Blue Shield Arizona",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "DME Guidelines",
        "notes": "DME rental covered up to purchase price. Policy search available at azblue.com. Prior authorization lookup tool available.",
        "source": "https://www.azblue.com/provider/resources/prior-authorization-and-medical-policies/search"
    },
    # Louisiana Blue (BCBS)
    {
        "name": "Louisiana Blue (BCBS Louisiana)",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Medical policies available at lablue.com/medicalpolicies. Check for specific HFCWO/E0469 coverage criteria.",
        "source": "https://www.lablue.com/medicalpolicies"
    },
    # BCBS Minnesota
    {
        "name": "Blue Cross Blue Shield Minnesota",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Evidence-based coverage criteria. Access medical and behavioral health policies at bluecrossmn.com provider portal.",
        "source": "https://www.bluecrossmn.com/providers/medical-management/medical-and-behavioral-health-policies"
    },
    # Hill-Rom/Baxter coverage reference
    {
        "name": "Medicare Coverage Reference (Hill-Rom)",
        "type": "Reference",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "LCD L33785",
        "notes": "Industry coverage criteria sheet for HFCWO. Lists common requirements: CF, bronchiectasis (CT confirmed), neuromuscular disease, failed standard therapies.",
        "source": "https://www.hillrom.com/content/dam/hillrom-aem/us/en/marketing/products/vest-apx-system/documents/US-FLC174-220030-v4-Vest-Monarch-Medicare-Coverage-Criteria-Sheet.pdf"
    },
    # Rocky Mountain Health Plans
    {
        "name": "Rocky Mountain Health Plans (Colorado)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "03/01/2025",
        "policy_num": "Airway Clearance Devices",
        "notes": "Policy addresses HFCWO, IPV devices. Codes A7021, A7025, A7026, E0469, E0481, E0483. Partners with UHC Community Plan in Colorado.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/exchange-policies/rmhp-medical-policies.html"
    },
    # Geisinger Health Plan
    {
        "name": "Geisinger Health Plan (Pennsylvania)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (some devices)",
        "not_med_necessary": "Yes (some devices)",
        "date": "Current",
        "policy_num": "MP045",
        "notes": "Covers E0469, E0480, E0481, E0482. Neuromuscular disease with respiratory weakness. Some devices NOT covered - insufficient evidence. Pre-cert required.",
        "source": "https://www.geisinger.org/-/media/OneGeisinger/Files/Policy%20PDFs/MP/1-50/MP045%20High%20Frequency%20Chest%20Percussion%20Vest.ashx"
    },
    # Univera Healthcare (NY)
    {
        "name": "Univera Healthcare (New York)",
        "type": "Regional",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Yes (replacements)",
        "date": "Current",
        "policy_num": "Policy 1.01.15",
        "notes": "E0469, E0481, E0482, E0483 covered. Referral must be from pulmonologist. Contract dependent. Replacement of functioning device NOT covered.",
        "source": "https://www.univerahealthcare.com/documents/d/global/uni-prv-airway-clearance-devices-1"
    },
    # Kaiser Permanente Washington (Group Health)
    {
        "name": "Kaiser Permanente Washington",
        "type": "HMO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Yes (neuromuscular)",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Criteria",
        "notes": "E0469 covered for CF (pulmonologist managed, compliance documented) or bronchiectasis. HFCWO for neuromuscular deficiency does NOT meet criteria.",
        "source": "https://wa-provider.kaiserpermanente.org/static/pdf/hosting/clinical/criteria/pdf/hfcwo.pdf"
    },
    # EmblemHealth (NY)
    {
        "name": "EmblemHealth (GHI/HIP New York)",
        "type": "Regional",
        "coverage": "Per Medicare Guidelines",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Follows CMS/LCD",
        "notes": "DME prior approval consistent with CMS/Medicare Coverage Guidelines. eviCore handles preauth for HIP/HIPIC members.",
        "source": "https://www.emblemhealth.com/providers/manual/durable-medical-equipment"
    },
    # Blue Shield California
    {
        "name": "Blue Shield of California",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Oscillatory Devices Policy",
        "notes": "Policy for Oscillatory Devices for CF and Other Respiratory Conditions. Access via blueshieldca.com/provider medical policies.",
        "source": "https://www.blueshieldca.com/en/provider/authorizations/policy-medical/list"
    },
    # Blue Cross Idaho
    {
        "name": "Blue Cross of Idaho",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "MP 9.01.501+",
        "notes": "Uses internal policies or InterQual criteria when no Medicare LCD/NCD exists. Check portal at providers.bcidaho.com.",
        "source": "https://providers.bcidaho.com/medical-management/medical-policies/medical-policies.page"
    },
    # BCBS Nebraska
    {
        "name": "Blue Cross Blue Shield Nebraska",
        "type": "BCBS",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Med Policy Blue",
        "notes": "Evidence-based Medical Policies developed by Medical Policy Committee. Search at medicalpolicy.nebraskablue.com.",
        "source": "https://medicalpolicy.nebraskablue.com/"
    },
    # Illinois Medicaid (HFS)
    {
        "name": "Illinois Medicaid (HFS)",
        "type": "Medicaid",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFS 2305B/2305C",
        "notes": "Questionnaire forms required for Airway Clearance Device (HFS 2305B) and Continued Rental (HFS 2305C).",
        "source": "https://hfs.illinois.gov/info/brochures-and-forms/medicalforms.html"
    },
    # Fallon Health (MA)
    {
        "name": "Fallon Health (Massachusetts)",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Medical policies available at fallonhealth.org/providers/criteria-policies-guidelines/medical-policies. Contact 1-800-868-5200.",
        "source": "https://fallonhealth.org/providers/criteria-policies-guidelines/medical-policies"
    },
    # MVP Health Care (NY/VT)
    {
        "name": "MVP Health Care (NY/VT)",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Provider Portal",
        "notes": "Medical policies require provider portal login. Policies reviewed annually. Access via mvphealthcare.com/providers.",
        "source": "https://www.mvphealthcare.com/-/media/project/mvp/healthcare/documents/provider/mvp-health-care-medical-policies.pdf"
    },
    # UPMC Health Plan
    {
        "name": "UPMC Health Plan (Pennsylvania)",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "See Portal",
        "notes": "Check UPMC Health Plan provider portal for specific HFCWO/E0469 coverage criteria. upmchealthplan.com.",
        "source": "https://www.upmchealthplan.com"
    },
    # Florida Medicaid (AHCA)
    {
        "name": "Florida Medicaid (AHCA)",
        "type": "Medicaid",
        "coverage": "Per State Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Fee Schedule",
        "notes": "Covered services via AHCA. Check Florida Medicaid fee schedules and coverage policies at ahca.myflorida.com.",
        "source": "https://ahca.myflorida.com/medicaid/medicaid-policy-quality-and-operations/medicaid-policy-and-quality/medicaid-policy/florida-medicaid-s-covered-services-and-hcbs-waivers"
    },
    # Sunshine Health (FL Medicaid MCO)
    {
        "name": "Sunshine Health (Florida Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Per State Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Centene Policy",
        "notes": "Florida Medicaid managed care through SMMC program. Check sunshinehealth.com for coverage details.",
        "source": "https://www.sunshinehealth.com/members/medicaid.html"
    },
    # Peach State Health Plan (GA)
    {
        "name": "Peach State Health Plan (Georgia)",
        "type": "Medicaid MCO",
        "coverage": "Per InterQual/Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "InterQual Criteria",
        "notes": "Uses InterQual criteria for DME. Check pshpgeorgia.com/providers for clinical payment policies.",
        "source": "https://www.pshpgeorgia.com/providers/resources/clinical-payment-policies.html"
    },
    # Superior HealthPlan (TX)
    {
        "name": "Superior HealthPlan (Texas)",
        "type": "Medicaid MCO",
        "coverage": "Per State Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "TX Medicaid",
        "notes": "Texas Medicaid services must be medically necessary. Check TMPPM at tmhp.com and superiorhealthplan.com.",
        "source": "https://www.superiorhealthplan.com/members/medicaid/benefits-services.html"
    },
    # ODS Health Plan (Oregon)
    {
        "name": "ODS Health Plan (Oregon)",
        "type": "Regional",
        "coverage": "Per Medical Policy",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "HFCWO Policy",
        "notes": "HFCWO medical necessity criteria available. References Oregon HERC guidance for coverage recommendations.",
        "source": "https://odscompanies.com/pdfs/med_criteria/HighFrequencyChestWallOscillationDevices.pdf"
    },
    # Medica - EXPLICITLY mentions E0469
    {
        "name": "Medica (MN/Regional)",
        "type": "Regional",
        "coverage": "NOT COVERED - Investigational",
        "prior_auth": "N/A",
        "investigational": "Yes",
        "not_med_necessary": "Yes",
        "date": "Current",
        "policy_num": "Volara OLE Policy",
        "notes": "E0469 EXPLICITLY listed. Oscillating lung expansion systems (Volara, BiWaze Clear) NOT COVERED - investigative/unproven. Insufficient peer-reviewed evidence.",
        "source": "https://partner.medica.com/-/media/documents/provider/coverage-policies/volara-oscillation-and-lung-expansion-cp.pdf"
    },
    # UnitedHealthcare UMR - EXPLICITLY mentions E0469
    {
        "name": "UMR (UnitedHealthcare TPA)",
        "type": "Commercial TPA",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed in applicable codes (A7021, A7025, A7026, E0469, E0481, E0483). HFCWO and IPV devices covered.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/commercial-policies/umr-medical-drug-policies.html"
    },
    # Surest (UnitedHealthcare) - EXPLICITLY mentions E0469
    {
        "name": "Surest (UnitedHealthcare)",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed in applicable codes. HFCWO and IPV devices policy same as UHC commercial.",
        "source": "https://ams-gateway.uhcprovider.com/en/policies-protocols/commercial-policies/surest-medical-drug-policies.html"
    },
    # UHC Individual Exchange - EXPLICITLY mentions E0469
    {
        "name": "UnitedHealthcare Individual Exchange",
        "type": "ACA Exchange",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "Airway Clearance Devices",
        "notes": "E0469 EXPLICITLY listed. Addresses HFCWO and IPV devices. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/policies-protocols/exchange-policies/exchanges-medical-drug-policies.html"
    },
    # UHC New Jersey Community Plan - EXPLICITLY mentions E0469
    {
        "name": "UnitedHealthcare New Jersey (Medicaid)",
        "type": "Medicaid MCO",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "11/01/2025",
        "policy_num": "NJ-CS Policy",
        "notes": "E0469 EXPLICITLY listed. Policy effective 11/01/2025. Codes: A7021, A7025, A7026, E0469, E0481, E0483.",
        "source": "https://www.uhcprovider.com/en/health-plans-by-state/new-jersey-health-plans/nj-comm-plan-home/nj-cp-policies/medicaid-community-state-policies-nj.html"
    },
    # Cigna - EXPLICITLY mentions E0469 in Policy 0069
    {
        "name": "Cigna (Policy 0069 - E0469)",
        "type": "Commercial",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "Policy 0069",
        "notes": "E0469 EXPLICITLY covered. Airway clearance devices for CF, bronchiectasis, neuromuscular disease. Precert via Electromed 952-758-9299.",
        "source": "https://static.cigna.com/assets/chcp/pdf/coveragePolicies/medical/mm_0069_coveragepositioncriteria_airway_clearance_devices.pdf"
    },
    # Humana Medicare Advantage - EXPLICITLY mentions E0469
    {
        "name": "Humana Medicare Advantage (E0469)",
        "type": "Medicare Advantage",
        "coverage": "Covered with Prior Auth",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "01/01/2026",
        "policy_num": "MA Prior Auth List",
        "notes": "E0469 EXPLICITLY requires prior authorization. 90-day grace period for new enrollees in active treatment. See CP2023011.",
        "source": "https://assets.humana.com/is/content/humana/FINAL_Medicare%20and%20DSNP%20Prior%20Authorization%20and%20Notification%20List%20-%201-1-2026pdf"
    },
    # Anthem Connecticut - EXPLICITLY mentions E0469
    {
        "name": "Anthem Blue Cross Connecticut (E0469)",
        "type": "BCBS",
        "coverage": "Covered with Criteria",
        "prior_auth": "Yes",
        "investigational": "Not Specified",
        "not_med_necessary": "Not Specified",
        "date": "Current",
        "policy_num": "CG-DME-43",
        "notes": "E0469 in High Frequency Chest Compression Devices for Airway Clearance policy. Prior auth required.",
        "source": "https://genhealth.ai/policy/anthem-bluecross-ct/6001b967-cg-dme-43-high-frequency-chest-compression-devices-for-airway-clearance"
    },
    # ABM Respiratory Care E0469 Reimbursement Reference
    {
        "name": "BiWaze Clear (ABM) Reimbursement Guide",
        "type": "Manufacturer Reference",
        "coverage": "Reference Only",
        "prior_auth": "Varies",
        "investigational": "Some Payers",
        "not_med_necessary": "Some Payers",
        "date": "10/01/2024",
        "policy_num": "E0469 + A7021",
        "notes": "E0469 effective 10/1/2024 for OLE therapy. A7021 for monthly disposables. Notes many payers consider investigational.",
        "source": "https://resources.abmrc.com/biwaze-clear-reimbursement-hcpcs-codes-e0469-a7021-for-ole-therapy"
    },
    # CMS Fee Schedule E0469 effective date
    {
        "name": "CMS DMEPOS Fee Schedule (E0469)",
        "type": "Medicare/CMS",
        "coverage": "Case-by-Case (No LCD)",
        "prior_auth": "N/A",
        "investigational": "No Determination",
        "not_med_necessary": "No Determination",
        "date": "10/01/2024",
        "policy_num": "No LCD/NCD",
        "notes": "E0469 added to DMEPOS fee schedule effective 10/1/2024. Category codes 01, 60. No specific LCD - claims reviewed individually.",
        "source": "https://www.cms.gov/files/document/r12835cp.pdf"
    },
]

# Write data
for row_num, payer in enumerate(payer_data, 2):
    ws.cell(row=row_num, column=1, value=payer["name"])
    ws.cell(row=row_num, column=2, value=payer["type"])

    coverage_cell = ws.cell(row=row_num, column=3, value=payer["coverage"])
    if "Investigational" in payer["coverage"]:
        coverage_cell.fill = investigational_fill
    elif "Not Covered" in payer["coverage"]:
        coverage_cell.fill = not_covered_fill
    elif "Case-by-Case" in payer["coverage"]:
        coverage_cell.fill = case_by_case_fill
    elif "Covered" in payer["coverage"]:
        coverage_cell.fill = covered_fill

    ws.cell(row=row_num, column=4, value=payer["prior_auth"])

    inv_cell = ws.cell(row=row_num, column=5, value=payer["investigational"])
    if payer["investigational"] == "Yes" or "Yes" in payer["investigational"]:
        inv_cell.fill = investigational_fill

    nmn_cell = ws.cell(row=row_num, column=6, value=payer["not_med_necessary"])
    if payer["not_med_necessary"] == "Yes" or "Yes" in payer["not_med_necessary"]:
        nmn_cell.fill = not_covered_fill

    ws.cell(row=row_num, column=7, value=payer["date"])
    ws.cell(row=row_num, column=8, value=payer["policy_num"])
    ws.cell(row=row_num, column=9, value=payer["notes"])
    ws.cell(row=row_num, column=10, value=payer["source"])

    # Apply borders and alignment
    for col in range(1, 11):
        cell = ws.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='top', wrap_text=True)

# Set column widths
column_widths = [30, 15, 25, 15, 22, 22, 18, 20, 60, 50]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Set row height for header
ws.row_dimensions[1].height = 30

# Freeze top row
ws.freeze_panes = "A2"

# Add summary sheet
ws2 = wb.create_sheet(title="Summary")
ws2.cell(row=1, column=1, value="E0469 Payer Coverage Analysis Summary")
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)

summary_data = [
    ("", ""),
    ("HCPCS Code:", "E0469"),
    ("Description:", "Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device"),
    ("Effective Date:", "October 1, 2024 (CMS introduced code)"),
    ("Related Code:", "A7021 (monthly disposables)"),
    ("", ""),
    ("COVERAGE SUMMARY:", ""),
    ("Total Payers Analyzed:", str(len(payer_data))),
    ("Covered with Criteria:", str(sum(1 for p in payer_data if "Covered" in p["coverage"] and "Investigational" not in p["coverage"]))),
    ("Investigational/Experimental:", str(sum(1 for p in payer_data if "Investigational" in p["coverage"] or p["investigational"].startswith("Yes")))),
    ("Case-by-Case Review:", str(sum(1 for p in payer_data if "Case-by-Case" in p["coverage"]))),
    ("Prior Auth Required:", str(sum(1 for p in payer_data if p["prior_auth"] == "Yes"))),
    ("", ""),
    ("KEY FINDINGS:", ""),
    ("1.", "Medicare has NO specific LCD or NCD for E0469 - claims reviewed individually"),
    ("2.", "Multiple BCBS plans consider OLE devices (Volara, BiWaze Clear, MetaNeb) INVESTIGATIONAL"),
    ("3.", "Coverage most commonly approved for: Cystic Fibrosis, Bronchiectasis, Neuromuscular diseases"),
    ("4.", "Most payers require prior authorization for airway clearance devices"),
    ("5.", "COPD and other conditions often excluded or considered investigational"),
    ("", ""),
    ("DEVICES COMMONLY NOTED AS INVESTIGATIONAL:", ""),
    ("- Volara System", "OLE 3-in-1 device"),
    ("- BiWaze Clear System", "OLE device"),
    ("- MetaNeb System", "CHFO device"),
    ("", ""),
    ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
]

for row_num, (label, value) in enumerate(summary_data, 2):
    ws2.cell(row=row_num, column=1, value=label)
    if label.endswith(":") or label.startswith("-"):
        ws2.cell(row=row_num, column=1).font = Font(bold=True)
    ws2.cell(row=row_num, column=2, value=value)

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 80

# Save workbook
output_path = "/Users/leahnoaeill/Downloads/MyClaude/data_export/E0469_Payer_Coverage_Analysis.xlsx"
wb.save(output_path)
print(f"Spreadsheet saved to: {output_path}")
print(f"Total payers analyzed: {len(payer_data)}")
