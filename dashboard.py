#!/usr/bin/env python3
"""
E0469 Payer Coverage Dashboard
Flask web application for viewing and managing payer coverage data.
"""

from flask import Flask, request, jsonify, render_template, Response
import psycopg2
from psycopg2.extras import RealDictCursor
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import requests
import re

app = Flask(__name__)

# Database configuration
DB_CONFIG = {
    "dbname": os.environ.get("DB_NAME", "e0469_analysis"),
    "user": os.environ.get("DB_USER", "postgres"),
    "host": os.environ.get("DB_HOST", "localhost"),
    "port": int(os.environ.get("DB_PORT", 5432))
}


def get_db_connection():
    """Get database connection with RealDictCursor."""
    return psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)


# Simplified coverage status categories (3 options)
COVERAGE_CATEGORIES = {
    "Covered": "#C6EFCE",          # Green
    "Not Covered": "#FFC7CE",       # Red
    "Prior-Auth Required": "#BDD7EE"  # Blue
}

# Map detailed statuses to simplified categories
def normalize_coverage_status(status):
    """Convert detailed coverage status to one of 3 simplified categories."""
    if not status:
        return "Prior-Auth Required"

    status_lower = status.lower()

    # Not Covered
    if 'not covered' in status_lower or 'non-reimbursable' in status_lower:
        return "Not Covered"

    # Covered (explicit coverage)
    if status_lower.startswith('covered'):
        return "Covered"

    # Everything else maps to Prior-Auth Required:
    # - Investigational, Partial, Case-by-Case, Prior Auth, Varies, Reference
    return "Prior-Auth Required"

# Legacy color mapping (for raw status display)
COVERAGE_COLORS = {
    "NOT COVERED": "#FFC7CE",
    "NOT COVERED - Experimental/Investigational": "#FFC7CE",
    "NOT COVERED - EIU Non-Reimbursable": "#FFC7CE",
    "Investigational": "#BDD7EE",
    "Investigational - Experimental": "#BDD7EE",
    "Partial - OLE Unproven": "#BDD7EE",
    "Partial - Limited Conditions": "#BDD7EE",
    "Partial - Some Investigational": "#BDD7EE",
    "Case-by-Case (No LCD)": "#BDD7EE",
    "Case Review - Prior Auth Needed": "#BDD7EE",
    "Prior Auth Required": "#BDD7EE",
    "Prior Auth Required (MA)": "#BDD7EE",
    "Prior Auth Required - Clinical Review": "#BDD7EE",
    "Covered with Criteria": "#C6EFCE",
    "Covered with Prior Auth": "#C6EFCE",
    "Covered with Limits": "#C6EFCE",
    "Covered - Fee Schedule": "#C6EFCE",
    "Covered - Per Fee Schedule": "#C6EFCE",
    "Covered - Rental Only": "#C6EFCE",
    "Varies - EIU or Clinical Review": "#BDD7EE",
    "Reference Only": "#BDD7EE",
    # Simplified categories
    "Covered": "#C6EFCE",
    "Not Covered": "#FFC7CE",
    "Prior-Auth Required": "#BDD7EE",
}


@app.route('/')
def dashboard():
    """Render main dashboard."""
    return render_template('dashboard.html')


@app.route('/api/payers')
def get_payers():
    """Get payers with filtering, sorting, and pagination."""
    # Parse query parameters
    name = request.args.get('name', '').strip()
    payer_type = request.args.get('payer_type', '').strip()
    coverage_status = request.args.get('coverage_status', '').strip()
    investigational = request.args.get('investigational', '').strip()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    sort_by = request.args.get('sort_by', 'name').strip()
    sort_dir = request.args.get('sort_dir', 'asc').strip().lower()

    # Validate sort parameters
    allowed_sort_fields = ['name', 'payer_type', 'coverage_status', 'prior_auth_required',
                           'investigational', 'policy_date', 'policy_number']
    if sort_by not in allowed_sort_fields:
        sort_by = 'name'
    if sort_dir not in ['asc', 'desc']:
        sort_dir = 'asc'

    # Build WHERE clause
    where_clauses = []
    params = []

    if name:
        where_clauses.append("p.name ILIKE %s")
        params.append(f"%{name}%")

    if payer_type:
        where_clauses.append("p.payer_type = %s")
        params.append(payer_type)

    if coverage_status:
        # Filter by coverage status (already normalized to 3 categories in DB)
        where_clauses.append("pp.coverage_status = %s")
        params.append(coverage_status)

    if investigational:
        if investigational == 'Yes':
            where_clauses.append("(pp.investigational LIKE %s OR pp.investigational = %s)")
            params.extend(['Yes%', 'Yes'])
        elif investigational == 'No':
            where_clauses.append("(pp.investigational = %s OR pp.investigational = %s)")
            params.extend(['No', 'No Determination'])

    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

    # Calculate offset
    offset = (page - 1) * per_page

    conn = get_db_connection()
    cur = conn.cursor()

    # Get total count
    cur.execute(f"""
        SELECT COUNT(DISTINCT p.id)
        FROM payers p
        LEFT JOIN payer_policies pp ON p.id = pp.payer_id
        WHERE {where_sql}
    """, params)
    total = cur.fetchone()['count']

    # Get payers with policies
    cur.execute(f"""
        SELECT
            p.id,
            p.name,
            p.payer_type,
            pp.coverage_status,
            pp.prior_auth_required,
            pp.investigational,
            pp.not_med_necessary,
            pp.policy_date,
            pp.policy_number,
            pp.notes,
            pp.source_url
        FROM payers p
        LEFT JOIN payer_policies pp ON p.id = pp.payer_id
        WHERE {where_sql}
        ORDER BY {sort_by} {sort_dir} NULLS LAST
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    payers = [dict(row) for row in cur.fetchall()]

    # Add color codes (coverage_status is already normalized in DB)
    for payer in payers:
        payer['coverage_category'] = payer['coverage_status']
        payer['color_code'] = COVERAGE_CATEGORIES.get(payer['coverage_status'], '#E2E8F0')

    conn.close()

    return jsonify({
        'payers': payers,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page if total > 0 else 1
    })


@app.route('/api/payers/<int:payer_id>')
def get_payer(payer_id):
    """Get single payer details."""
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.id,
            p.name,
            p.payer_type,
            pp.coverage_status,
            pp.prior_auth_required,
            pp.investigational,
            pp.not_med_necessary,
            pp.policy_date,
            pp.policy_number,
            pp.notes,
            pp.source_url
        FROM payers p
        LEFT JOIN payer_policies pp ON p.id = pp.payer_id
        WHERE p.id = %s
    """, [payer_id])

    payer = cur.fetchone()
    conn.close()

    if payer:
        payer = dict(payer)
        payer['coverage_category'] = payer['coverage_status']
        payer['color_code'] = COVERAGE_CATEGORIES.get(payer['coverage_status'], '#E2E8F0')
        return jsonify(payer)
    else:
        return jsonify({'error': 'Payer not found'}), 404


@app.route('/api/payers/<int:payer_id>', methods=['PUT'])
def update_payer(payer_id):
    """Update payer policy information."""
    data = request.get_json()

    conn = get_db_connection()
    cur = conn.cursor()

    # Update payer_policies
    update_fields = []
    params = []

    if 'coverage_status' in data:
        update_fields.append("coverage_status = %s")
        params.append(data['coverage_status'])

    if 'prior_auth_required' in data:
        update_fields.append("prior_auth_required = %s")
        params.append(data['prior_auth_required'])

    if 'investigational' in data:
        update_fields.append("investigational = %s")
        params.append(data['investigational'])

    if 'notes' in data:
        update_fields.append("notes = %s")
        params.append(data['notes'])

    if update_fields:
        params.append(payer_id)
        cur.execute(f"""
            UPDATE payer_policies
            SET {', '.join(update_fields)}, updated_at = CURRENT_TIMESTAMP
            WHERE payer_id = %s
        """, params)
        conn.commit()

    conn.close()

    return jsonify({'success': True})


@app.route('/api/coverage-statuses')
def get_coverage_statuses():
    """Get the 3 simplified coverage categories."""
    # Return only the 3 simplified categories
    return jsonify(["Covered", "Prior-Auth Required", "Not Covered"])


@app.route('/api/payer-types')
def get_payer_types():
    """Get distinct payer types."""
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT payer_type
        FROM payers
        WHERE payer_type IS NOT NULL
        ORDER BY payer_type
    """)

    types = [row['payer_type'] for row in cur.fetchall()]
    conn.close()

    return jsonify(types)


@app.route('/api/aggregates')
def get_aggregates():
    """Get summary statistics."""
    conn = get_db_connection()
    cur = conn.cursor()

    # Total payers
    cur.execute("SELECT COUNT(*) as count FROM payers")
    total_payers = cur.fetchone()['count']

    # Total searched (no policy)
    cur.execute("SELECT COUNT(*) as count FROM searched_payers")
    total_searched = cur.fetchone()['count']

    # Coverage status counts
    cur.execute("""
        SELECT coverage_status, COUNT(*) as count
        FROM payer_policies
        GROUP BY coverage_status
        ORDER BY count DESC
    """)
    coverage_counts = [dict(row) for row in cur.fetchall()]

    # Payer type counts
    cur.execute("""
        SELECT payer_type, COUNT(*) as count
        FROM payers
        GROUP BY payer_type
        ORDER BY count DESC
    """)
    type_counts = [dict(row) for row in cur.fetchall()]

    # Investigational status
    cur.execute("""
        SELECT
            CASE
                WHEN investigational LIKE 'Yes%' OR investigational = 'Yes' THEN 'Investigational'
                WHEN investigational = 'No' OR investigational = 'No Determination' THEN 'Not Investigational'
                ELSE 'Not Specified'
            END as status,
            COUNT(*) as count
        FROM payer_policies
        GROUP BY status
        ORDER BY count DESC
    """)
    investigational_counts = [dict(row) for row in cur.fetchall()]

    # Coverage summary (coverage_status is already normalized to 3 categories)
    cur.execute("""
        SELECT coverage_status as category, COUNT(*) as count
        FROM payer_policies
        GROUP BY coverage_status
        ORDER BY
            CASE coverage_status
                WHEN 'Covered' THEN 1
                WHEN 'Prior-Auth Required' THEN 2
                WHEN 'Not Covered' THEN 3
            END
    """)
    summary_counts = [dict(row) for row in cur.fetchall()]

    conn.close()

    return jsonify({
        'total_payers': total_payers,
        'total_searched': total_searched,
        'coverage_counts': coverage_counts,
        'type_counts': type_counts,
        'investigational_counts': investigational_counts,
        'summary_counts': summary_counts
    })


@app.route('/api/searched-payers')
def get_searched_payers():
    """Get payers that were searched but no E0469 policy found."""
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    payer_type = request.args.get('payer_type', '').strip()

    offset = (page - 1) * per_page

    conn = get_db_connection()
    cur = conn.cursor()

    where_sql = "1=1"
    params = []

    if payer_type:
        where_sql = "payer_type = %s"
        params.append(payer_type)

    # Get total
    cur.execute(f"SELECT COUNT(*) as count FROM searched_payers WHERE {where_sql}", params)
    total = cur.fetchone()['count']

    # Get payers
    cur.execute(f"""
        SELECT id, name, payer_type, notes, date_searched
        FROM searched_payers
        WHERE {where_sql}
        ORDER BY name
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    payers = [dict(row) for row in cur.fetchall()]
    conn.close()

    return jsonify({
        'payers': payers,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page if total > 0 else 1
    })


@app.route('/api/export')
def export_excel():
    """Export payer data to Excel."""
    # Get filter parameters
    name = request.args.get('name', '').strip()
    payer_type = request.args.get('payer_type', '').strip()
    coverage_status = request.args.get('coverage_status', '').strip()

    conn = get_db_connection()
    cur = conn.cursor()

    # Build WHERE clause
    where_clauses = []
    params = []

    if name:
        where_clauses.append("p.name ILIKE %s")
        params.append(f"%{name}%")

    if payer_type:
        where_clauses.append("p.payer_type = %s")
        params.append(payer_type)

    if coverage_status:
        where_clauses.append("pp.coverage_status = %s")
        params.append(coverage_status)

    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

    # Get payers with policies
    cur.execute(f"""
        SELECT
            p.name,
            p.payer_type,
            pp.coverage_status,
            pp.prior_auth_required,
            pp.investigational,
            pp.not_med_necessary,
            pp.policy_date,
            pp.policy_number,
            pp.notes,
            pp.source_url
        FROM payers p
        LEFT JOIN payer_policies pp ON p.id = pp.payer_id
        WHERE {where_sql}
        ORDER BY p.name
    """, params)

    payers = cur.fetchall()

    # Get searched payers
    cur.execute("""
        SELECT name, payer_type, notes, date_searched
        FROM searched_payers
        ORDER BY name
    """)
    searched = cur.fetchall()

    conn.close()

    # Create workbook
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1: Payers with E0469 Policies
    ws1 = wb.active
    ws1.title = "E0469 Payer Policies"

    headers1 = ["Payer Name", "Payer Type", "Coverage Status", "Prior Auth Required",
                "Investigational", "Not Med Necessary", "Policy Date", "Policy Number",
                "Notes", "Source URL"]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for row_num, payer in enumerate(payers, 2):
        ws1.cell(row=row_num, column=1, value=payer['name'])
        ws1.cell(row=row_num, column=2, value=payer['payer_type'])

        coverage_cell = ws1.cell(row=row_num, column=3, value=payer['coverage_status'])
        color = COVERAGE_COLORS.get(payer['coverage_status'], '#E2E8F0').replace('#', '')
        coverage_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws1.cell(row=row_num, column=4, value=payer['prior_auth_required'])
        ws1.cell(row=row_num, column=5, value=payer['investigational'])
        ws1.cell(row=row_num, column=6, value=payer['not_med_necessary'])
        ws1.cell(row=row_num, column=7, value=payer['policy_date'])
        ws1.cell(row=row_num, column=8, value=payer['policy_number'])
        ws1.cell(row=row_num, column=9, value=payer['notes'])
        ws1.cell(row=row_num, column=10, value=payer['source_url'])

        for col in range(1, 11):
            ws1.cell(row=row_num, column=col).border = thin_border
            ws1.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)

    # Set column widths
    widths1 = [35, 18, 30, 20, 20, 20, 15, 25, 60, 50]
    for col, width in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(col)].width = width

    ws1.freeze_panes = "A2"

    # Sheet 2: Searched Payers (No E0469)
    ws2 = wb.create_sheet("Searched - No E0469")

    headers2 = ["Payer Name", "Payer Type", "Notes", "Date Searched"]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    for row_num, payer in enumerate(searched, 2):
        ws2.cell(row=row_num, column=1, value=payer['name'])
        ws2.cell(row=row_num, column=2, value=payer['payer_type'])
        ws2.cell(row=row_num, column=3, value=payer['notes'])
        ws2.cell(row=row_num, column=4, value=str(payer['date_searched']) if payer['date_searched'] else '')

        for col in range(1, 5):
            ws2.cell(row=row_num, column=col).border = thin_border
            ws2.cell(row=row_num, column=col).alignment = Alignment(vertical='top', wrap_text=True)

    widths2 = [40, 25, 60, 15]
    for col, width in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    ws2.freeze_panes = "A2"

    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="E0469 Payer Coverage Analysis Summary")
    ws3.cell(row=1, column=1).font = Font(bold=True, size=14)

    summary_data = [
        ("", ""),
        ("HCPCS Code:", "E0469"),
        ("Description:", "Lung expansion airway clearance, continuous high frequency oscillation, and nebulization device"),
        ("Effective Date:", "October 1, 2024"),
        ("", ""),
        ("Total Payers with Explicit E0469 Policies:", str(len(payers))),
        ("Searched Payers (No E0469 Found):", str(len(searched))),
        ("", ""),
        ("Report Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for row_num, (label, value) in enumerate(summary_data, 3):
        ws3.cell(row=row_num, column=1, value=label)
        if label.endswith(":"):
            ws3.cell(row=row_num, column=1).font = Font(bold=True)
        ws3.cell(row=row_num, column=2, value=value)

    ws3.column_dimensions['A'].width = 45
    ws3.column_dimensions['B'].width = 80

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"E0469_Payer_Coverage_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )


@app.route('/api/payers', methods=['POST'])
def add_payer():
    """Add a new payer with policy."""
    data = request.get_json()

    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Payer name is required'}), 400

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # Check if payer already exists
        cur.execute("SELECT id FROM payers WHERE name = %s", [name])
        existing = cur.fetchone()
        if existing:
            return jsonify({'error': f'Payer "{name}" already exists'}), 400

        # Insert payer
        cur.execute("""
            INSERT INTO payers (name, payer_type)
            VALUES (%s, %s)
            RETURNING id
        """, (name, data.get('payer_type', '')))
        payer_id = cur.fetchone()['id']

        # Insert policy if coverage status provided
        if data.get('coverage_status') or data.get('source_url') or data.get('notes'):
            cur.execute("""
                INSERT INTO payer_policies (
                    payer_id, coverage_status, prior_auth_required,
                    investigational, notes, source_url
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                payer_id,
                data.get('coverage_status', ''),
                data.get('prior_auth_required', ''),
                data.get('investigational', ''),
                data.get('notes', ''),
                data.get('source_url', '')
            ))

        conn.commit()
        return jsonify({'success': True, 'id': payer_id})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/web-search', methods=['POST'])
def web_search():
    """Search the web for E0469 payer policies."""
    data = request.get_json()
    query = data.get('query', '').strip()

    if not query:
        return jsonify({'error': 'Search query is required'}), 400

    # Build search query for E0469 policies
    search_query = f'"{query}" "E0469" policy site:.gov OR site:.com OR site:.org filetype:pdf OR medical policy'

    try:
        # Use DuckDuckGo HTML search (no API key required)
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36'
        }

        # Search DuckDuckGo
        search_url = f"https://html.duckduckgo.com/html/?q={requests.utils.quote(search_query)}"
        response = requests.get(search_url, headers=headers, timeout=10)

        results = []

        if response.status_code == 200:
            # Parse results from HTML
            # Extract URLs and titles from DuckDuckGo results
            html = response.text

            # Find result links - DuckDuckGo uses class="result__a"
            pattern = r'<a[^>]*class="result__a"[^>]*href="([^"]*)"[^>]*>([^<]*)</a>'
            matches = re.findall(pattern, html, re.IGNORECASE)

            for url, title in matches[:10]:  # Limit to 10 results
                # Clean up URL (DuckDuckGo wraps URLs)
                if 'uddg=' in url:
                    url_match = re.search(r'uddg=([^&]+)', url)
                    if url_match:
                        url = requests.utils.unquote(url_match.group(1))

                # Filter for likely policy documents
                if any(x in url.lower() for x in ['policy', 'medical', 'coverage', '.pdf', 'provider']):
                    results.append({
                        'title': title.strip(),
                        'url': url
                    })

        # Also search Google (backup)
        if len(results) < 5:
            google_url = f"https://www.google.com/search?q={requests.utils.quote(search_query)}"
            try:
                g_response = requests.get(google_url, headers=headers, timeout=10)
                if g_response.status_code == 200:
                    # Extract URLs from Google results
                    g_pattern = r'<a[^>]*href="/url\?q=([^"&]+)[^"]*"'
                    g_matches = re.findall(g_pattern, g_response.text)
                    for url in g_matches[:5]:
                        url = requests.utils.unquote(url)
                        if not any(x in url for x in ['google.com', 'youtube.com', 'facebook.com']):
                            if url not in [r['url'] for r in results]:
                                results.append({
                                    'title': url.split('/')[-1][:50] or 'Policy Document',
                                    'url': url
                                })
            except:
                pass  # Google search is optional fallback

        return jsonify({
            'query': query,
            'results': results[:10]
        })

    except requests.Timeout:
        return jsonify({'error': 'Search timed out. Please try again.'}), 504
    except Exception as e:
        return jsonify({'error': f'Search failed: {str(e)}'}), 500


if __name__ == '__main__':
    print("Starting E0469 Payer Coverage Dashboard...")
    print(f"Database: {DB_CONFIG['dbname']} on {DB_CONFIG['host']}:{DB_CONFIG['port']}")
    print("Dashboard URL: http://localhost:5002")
    app.run(host='0.0.0.0', port=5002, debug=True)
